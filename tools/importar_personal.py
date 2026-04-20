"""
Importador del catálogo PERSONAL.

Fuentes (en orden de prioridad):
    1) BASE VACACIONES 2026.xlsx  hoja PERSONAL      (fuente canónica del tipo S/C)
    2) BASE GENERAL.xlsx          hoja 'PERSONAL '   (categoría/horario/situación médica)

Procedimiento seguro:
    --staging      Lee ambas hojas y llena import_personal_raw (NO toca personal).
    --dry-run      Simula upsert hacia personal reportando acciones.
    --importar     Procesa staging y hace upsert real en personal.

Matching (definido también en PersonalModel PHP):
    1. por numero_empleado cuando existe
    2. si no, por nombre_normalizado (upper, unaccent, sin espacios duplicados)
    3. si hay ambigüedad (mismo nombre pero num_empleado distinto en BD),
       la fila queda marcada estado_revision='PENDIENTE_REVISION'.

Variables de entorno (o valores por defecto locales):
    PGHOST=localhost PGPORT=5432 PGUSER=postgres
    PGPASSWORD=admin  PGDATABASE=respaldooficios
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import unicodedata
from datetime import datetime, date
from typing import Any, Optional

try:
    import psycopg2
    from psycopg2.extras import execute_batch, Json
except ImportError:
    print("ERROR: falta psycopg2. Instalar con:  pip install psycopg2-binary")
    sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. Instalar con:  pip install openpyxl")
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

XLSX_VAC = os.path.join(ROOT, "BASE VACACIONES 2026.xlsx")
XLSX_GEN = os.path.join(ROOT, "BASE GENERAL.xlsx")


def pg_connect():
    return psycopg2.connect(
        host     = os.environ.get("PGHOST",     "localhost"),
        port     = int(os.environ.get("PGPORT", "5432")),
        dbname   = os.environ.get("PGDATABASE", "respaldooficios"),
        user     = os.environ.get("PGUSER",     "postgres"),
        password = os.environ.get("PGPASSWORD", "admin"),
    )


# ---------------------------------------------------------------------
# utilidades
# ---------------------------------------------------------------------
def norm_value(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def normalizar_nombre(s: Optional[str]) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = " ".join(s.upper().split())
    return s


def to_jsonable(row: tuple) -> list:
    out = []
    for v in row:
        if isinstance(v, (datetime, date)):
            out.append(v.strftime("%Y-%m-%d"))
        else:
            out.append(v)
    return out


# ---------------------------------------------------------------------
# FASE 1 - STAGING
# ---------------------------------------------------------------------
def cargar_staging(conn) -> int:
    """Vuelca ambas hojas a import_personal_raw. Hace TRUNCATE previo."""
    rows_all = []

    # --- BASE VACACIONES 2026.xlsx : PERSONAL ---
    if os.path.exists(XLSX_VAC):
        wb = load_workbook(XLSX_VAC, data_only=True, read_only=True)
        if "PERSONAL" in wb.sheetnames:
            ws = wb["PERSONAL"]
            fuente = "BASE VACACIONES 2026.xlsx!PERSONAL"
            rows = list(ws.iter_rows(values_only=True))
            for idx, r in enumerate(rows[1:], start=2):
                if not r or all(v is None for v in r):
                    continue
                # cols: CONS, NO. EMPLEADO, NOMBRE, SINDICALIZADO, CONFIANZA
                cons, num, nombre, sind, conf = (r + (None,) * 5)[:5]
                tipo = None
                if (sind or "").strip().upper() == "S":
                    tipo = "S"
                elif (conf or "").strip().upper() == "C":
                    tipo = "C"
                rows_all.append((
                    fuente, idx,
                    norm_value(num), norm_value(nombre), tipo,
                    None, None, None,       # categoria, horario, sit_med
                    None, None, None, None, None,  # curp, rfc, issemym, estudios, carrera
                    Json(to_jsonable(r)),
                ))
        wb.close()

    # --- BASE GENERAL.xlsx : 'PERSONAL ' ---
    if os.path.exists(XLSX_GEN):
        wb = load_workbook(XLSX_GEN, data_only=True, read_only=True)
        hoja = None
        for s in wb.sheetnames:
            if s.strip().upper() == "PERSONAL":
                hoja = s; break
        if hoja:
            ws = wb[hoja]
            fuente = f"BASE GENERAL.xlsx!{hoja.strip()}"
            rows = list(ws.iter_rows(values_only=True))
            for idx, r in enumerate(rows[1:], start=2):
                if not r or all(v is None for v in r):
                    continue
                # cols: CONS, NO. EMPLEADO, NOMBRE, NOMINA, CATEGORIA, HORARIO,
                #       SITUACION MEDICA, CURP, RFC, CLAVE ISSEMYM, ESTUDIOS, CARRERA
                (cons, num, nombre, nomina, cat, horario, sit_med,
                 curp, rfc, issemym, estudios, carrera) = (r + (None,) * 12)[:12]
                tipo = None
                n = (nomina or "").strip().upper()
                if n == "S":
                    tipo = "S"
                elif n == "C":
                    tipo = "C"
                rows_all.append((
                    fuente, idx,
                    norm_value(num), norm_value(nombre), tipo,
                    norm_value(cat), norm_value(horario), norm_value(sit_med),
                    norm_value(curp), norm_value(rfc), norm_value(issemym),
                    norm_value(estudios), norm_value(carrera),
                    Json(to_jsonable(r)),
                ))
        wb.close()

    if not rows_all:
        print("No hay filas para cargar en staging.")
        return 0

    with conn.cursor() as cur:
        cur.execute("TRUNCATE import_personal_raw RESTART IDENTITY")
        execute_batch(
            cur,
            """INSERT INTO import_personal_raw
                   (fuente, fila_excel, raw_numero, raw_nombre, raw_tipo,
                    raw_categoria, raw_horario, raw_sit_medica,
                    raw_curp, raw_rfc, raw_clave_issemym,
                    raw_grado_estudios, raw_carrera, raw_data)
               VALUES (%s, %s, %s, %s, %s,
                       %s, %s, %s,
                       %s, %s, %s,
                       %s, %s, %s)""",
            rows_all,
            page_size=100,
        )
        conn.commit()

    print(f"Staging import_personal_raw: {len(rows_all)} filas cargadas.")
    return len(rows_all)


# ---------------------------------------------------------------------
# FASE 2 - IMPORTAR STAGING -> personal
# ---------------------------------------------------------------------
def procesar(conn, dry_run: bool = False) -> None:
    with conn.cursor() as cur:
        cur.execute("SELECT id, clave FROM tipos_personal")
        mapa_tipo = {clave: tid for (tid, clave) in cur.fetchall()}
        if "SINDICALIZADO" not in mapa_tipo or "CONFIANZA" not in mapa_tipo:
            print("ERROR: catálogo tipos_personal incompleto. Ejecute 05_personal.sql primero.")
            return

        # Traer staging pendiente ordenada por fuente (vacaciones primero: fuente canónica)
        cur.execute(
            """SELECT id, fuente, fila_excel, raw_numero, raw_nombre, raw_tipo,
                      raw_categoria, raw_horario, raw_sit_medica,
                      raw_curp, raw_rfc, raw_clave_issemym,
                      raw_grado_estudios, raw_carrera
                 FROM import_personal_raw
                WHERE procesado = FALSE
                ORDER BY
                   CASE WHEN fuente LIKE 'BASE VACACIONES%%' THEN 0 ELSE 1 END,
                   fila_excel"""
        )
        filas = cur.fetchall()

    ok = creados = actualizados = omitidos = errores = revision = 0
    procesados_dict = {}  # cache id_staging -> resultado

    for (sid, fuente, fila, num, nombre, tipo_raw,
         cat, horario, sitmed, curp, rfc, issemym, estudios, carrera) in filas:
        try:
            if not nombre:
                raise ValueError(f"Fila {fila}: nombre vacío.")

            # Tipo: si viene del staging usarlo; si no, si ya existe en BD por num, respetar BD; si no => ERROR
            tipo_clave = None
            if tipo_raw == "S":
                tipo_clave = "SINDICALIZADO"
            elif tipo_raw == "C":
                tipo_clave = "CONFIANZA"
            else:
                with conn.cursor() as c2:
                    if num:
                        c2.execute("SELECT tp.clave FROM personal p JOIN tipos_personal tp ON tp.id=p.tipo_personal_id WHERE p.numero_empleado=%s", (num,))
                        row = c2.fetchone()
                        if row: tipo_clave = row[0]
                    if not tipo_clave:
                        raise ValueError(f"Fila {fila}: no se puede determinar tipo (S/C) para '{nombre}'.")
            tipo_id = mapa_tipo[tipo_clave]

            norm = normalizar_nombre(nombre)

            # Matching
            with conn.cursor() as c:
                existe_id = None
                if num:
                    c.execute("SELECT id FROM personal WHERE numero_empleado=%s LIMIT 1", (num,))
                    row = c.fetchone()
                    if row: existe_id = row[0]

                if not existe_id:
                    c.execute("SELECT id, numero_empleado FROM personal WHERE nombre_normalizado=%s LIMIT 1", (norm,))
                    row = c.fetchone()
                    if row:
                        existe_id, num_bd = row
                        if num_bd and num and num_bd != num:
                            # Ambigüedad: marcar revisión y NO tocar
                            revision += 1
                            with conn.cursor() as cu:
                                cu.execute(
                                    """UPDATE import_personal_raw
                                          SET estado_revision='PENDIENTE_REVISION',
                                              error_importacion=%s
                                        WHERE id=%s""",
                                    (f"Mismo nombre pero num_empleado distinto (BD={num_bd}, Excel={num})", sid)
                                )
                                conn.commit()
                            continue

                if dry_run:
                    accion = "ACTUALIZAR" if existe_id else "CREAR"
                    print(f"  [DRY] {fila:>4}  {accion:<10} num={num}  nom='{nombre[:40]}'  tipo={tipo_clave}")
                    ok += 1
                    continue

                if existe_id:
                    # Actualización parcial: COALESCE para no pisar datos buenos, y actualizar num si BD no lo tiene.
                    cu = conn.cursor()
                    cu.execute(
                        """UPDATE personal
                              SET numero_empleado   = COALESCE(numero_empleado, %s),
                                  tipo_personal_id  = COALESCE(tipo_personal_id, %s),
                                  categoria         = COALESCE(categoria, %s),
                                  horario           = COALESCE(horario, %s),
                                  situacion_medica  = COALESCE(situacion_medica, %s),
                                  curp              = COALESCE(curp, %s),
                                  rfc               = COALESCE(rfc, %s),
                                  clave_issemym     = COALESCE(clave_issemym, %s),
                                  grado_estudios    = COALESCE(grado_estudios, %s),
                                  carrera           = COALESCE(carrera, %s)
                            WHERE id = %s""",
                        (num, tipo_id, cat, horario, sitmed,
                         curp, rfc, issemym, estudios, carrera, existe_id)
                    )
                    pid = existe_id
                    actualizados += 1
                    accion = "ACTUALIZADO"
                else:
                    cu = conn.cursor()
                    cu.execute(
                        """INSERT INTO personal
                           (numero_empleado, nombre_completo, nombre_normalizado, tipo_personal_id,
                            categoria, horario, situacion_medica, curp, rfc,
                            clave_issemym, grado_estudios, carrera)
                           VALUES (%s,%s,%s,%s, %s,%s,%s,%s,%s, %s,%s,%s)
                           RETURNING id""",
                        (num, nombre.strip(), norm, tipo_id,
                         cat, horario, sitmed, curp, rfc,
                         issemym, estudios, carrera)
                    )
                    pid = cu.fetchone()[0]
                    creados += 1
                    accion = "CREADO"

                # Marcar staging
                cu.execute(
                    """UPDATE import_personal_raw
                          SET procesado=TRUE, personal_id=%s, accion=%s,
                              estado_revision='OK', error_importacion=NULL
                        WHERE id=%s""",
                    (pid, accion, sid)
                )
                conn.commit()
                ok += 1
                procesados_dict[sid] = (pid, accion)

        except Exception as e:
            errores += 1
            conn.rollback()
            with conn.cursor() as cerr:
                cerr.execute(
                    """UPDATE import_personal_raw
                          SET estado_revision='ERROR', error_importacion=%s
                        WHERE id=%s""",
                    (str(e), sid)
                )
                conn.commit()
            print(f"  [ERR fila {fila}] {e}")

    print("\n--- Resultados personal ---")
    print(f"OK={ok}  creados={creados}  actualizados={actualizados}  "
          f"omitidos={omitidos}  revision={revision}  errores={errores}  "
          f"total_staging={len(filas)}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--staging",   action="store_true", help="Llena SOLO staging (seguro, NO toca personal).")
    ap.add_argument("--dry-run",   action="store_true", help="Simula importación, no escribe en personal.")
    ap.add_argument("--importar",  action="store_true", help="Importa staging -> personal.")
    args = ap.parse_args()

    if not (args.staging or args.dry_run or args.importar):
        ap.print_help(); sys.exit(0)

    conn = pg_connect()
    try:
        if args.staging:
            cargar_staging(conn)
        if args.dry_run:
            procesar(conn, dry_run=True)
        if args.importar:
            procesar(conn, dry_run=False)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
