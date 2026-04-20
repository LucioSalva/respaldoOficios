"""
Script de inspección (read-only) para listar hojas y primeras filas de los
Excel de origen. Se usa para entender la estructura real antes de diseñar
el esquema normalizado de personal/vacaciones/folios interno sub.

Uso:
    python tools/_inspect_excels.py
"""
from __future__ import annotations
import os
import sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FILES = [
    "BASE GENERAL.xlsx",
    "BASE VACACIONES 2026.xlsx",
    "BASE LUCIO.xlsx",
]

MAX_ROWS_PREVIEW = 5
MAX_COL_PREVIEW  = 20

def inspect(path: str) -> None:
    print("\n" + "=" * 78)
    print(f"ARCHIVO: {path}")
    print("=" * 78)
    if not os.path.exists(path):
        print("  [NO ENCONTRADO]")
        return
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n-- Hoja: {sheet!r}  (max_row={ws.max_row}, max_col={ws.max_column})")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("   (vacía)")
            continue
        # header
        header = rows[0][:MAX_COL_PREVIEW]
        print("   Header:")
        for i, h in enumerate(header, start=1):
            print(f"     col {i:>2}: {h!r}")
        # preview
        print("   Muestra de filas:")
        for idx, r in enumerate(rows[1:MAX_ROWS_PREVIEW+1], start=2):
            vals = r[:MAX_COL_PREVIEW]
            print(f"     fila {idx}: {vals}")
    wb.close()

def main() -> None:
    for f in FILES:
        inspect(os.path.join(ROOT, f))

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
