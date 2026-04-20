"""
Inspeccion profunda de hojas PERSONAL / VACACIONES / FOLIOS INTERNO SUB.
Busca: duplicados, filas vacias, calidad, conteos.
"""
from __future__ import annotations
import os
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def norm_name(s):
    if s is None: return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = " ".join(s.upper().split())
    return s.strip()

def val(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v
    s = str(v).strip()
    return s if s else None

def dump_section(title):
    print("\n" + "=" * 78)
    print(title)
    print("=" * 78)

# -----------------------------------------------------------
# 1) PERSONAL de BASE VACACIONES 2026.xlsx  (fuente canonica)
# -----------------------------------------------------------
def analizar_personal_vac():
    path = os.path.join(ROOT, "BASE VACACIONES 2026.xlsx")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["PERSONAL"]
    rows = list(ws.iter_rows(values_only=True))
    dump_section("PERSONAL (BASE VACACIONES 2026.xlsx) - Fuente canonica de tipo")

    filas = []
    for i, r in enumerate(rows[1:], start=2):
        if all(v is None for v in r): continue
        cons, num_emp, nombre, sind, conf = (r + (None,)*5)[:5]
        filas.append({
            "fila": i, "cons": cons, "num_emp": num_emp, "nombre": nombre,
            "sind": sind, "conf": conf,
        })
    print(f"Total filas con datos: {len(filas)}")

    # tipo
    por_tipo = Counter()
    sin_tipo = []
    ambos = []
    for f in filas:
        s = (f["sind"] or "").strip().upper()
        c = (f["conf"] or "").strip().upper()
        if s and c:
            ambos.append(f["fila"])
            por_tipo["AMBOS"] += 1
        elif s == "S":
            por_tipo["SINDICALIZADO"] += 1
        elif c == "C":
            por_tipo["CONFIANZA"] += 1
        else:
            sin_tipo.append(f["fila"])
            por_tipo["SIN_TIPO"] += 1
    print(f"Distribucion tipo: {dict(por_tipo)}")
    if sin_tipo: print(f"  filas sin tipo: {sin_tipo}")
    if ambos:    print(f"  filas con ambos: {ambos}")

    # duplicados por num_emp
    nums = [f["num_emp"] for f in filas if f["num_emp"] is not None]
    dup_nums = [n for n, c in Counter(nums).items() if c > 1]
    print(f"Numeros empleado duplicados: {dup_nums}")

    # duplicados por nombre normalizado
    nombres = [norm_name(f["nombre"]) for f in filas if f["nombre"]]
    dup_nom = [n for n, c in Counter(nombres).items() if c > 1]
    print(f"Nombres duplicados (normalizados): {dup_nom}")

    # num_emp nulo
    sin_num = [f["fila"] for f in filas if not f["num_emp"]]
    print(f"Filas sin numero_empleado: {sin_num}")

    # dump
    print("\nListado completo:")
    for f in filas:
        t = "SIND" if (f["sind"] or "").strip().upper()=="S" else \
            "CONF" if (f["conf"] or "").strip().upper()=="C" else "?"
        print(f"  fila {f['fila']:>3} | {f['num_emp']!s:<8} | {t:<4} | {f['nombre']}")
    wb.close()
    return filas

# -----------------------------------------------------------
# 2) PERSONAL de BASE GENERAL.xlsx (tiene CURP/RFC/CATEGORIA)
# -----------------------------------------------------------
def analizar_personal_gen():
    path = os.path.join(ROOT, "BASE GENERAL.xlsx")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["PERSONAL "]
    rows = list(ws.iter_rows(values_only=True))
    dump_section("PERSONAL (BASE GENERAL.xlsx) - con categoria/horario")

    filas = []
    for i, r in enumerate(rows[1:], start=2):
        if all(v is None for v in r): continue
        cons, num_emp, nombre, nomina, cat, horario, sit_med, curp, rfc, clave_issemym, estudios, carrera = (r + (None,)*12)[:12]
        filas.append({
            "fila": i, "num_emp": num_emp, "nombre": nombre,
            "nomina": nomina, "categoria": cat, "horario": horario,
            "sit_med": sit_med, "curp": curp, "rfc": rfc,
            "clave_issemym": clave_issemym, "estudios": estudios, "carrera": carrera,
        })
    print(f"Total filas con datos: {len(filas)}")

    por_nomina = Counter((f["nomina"] or "").strip().upper() for f in filas)
    print(f"Distribucion NOMINA: {dict(por_nomina)}")  # 'S'/'C'

    nums = [f["num_emp"] for f in filas if f["num_emp"] is not None]
    dup_nums = [n for n, c in Counter(nums).items() if c > 1]
    print(f"Numeros empleado duplicados: {dup_nums}")

    # campos faltantes
    sin_curp = sum(1 for f in filas if not f["curp"])
    sin_rfc  = sum(1 for f in filas if not f["rfc"])
    sin_cat  = sum(1 for f in filas if not f["categoria"])
    sin_hor  = sum(1 for f in filas if not f["horario"])
    print(f"Sin CURP: {sin_curp}/{len(filas)}")
    print(f"Sin RFC:  {sin_rfc}/{len(filas)}")
    print(f"Sin categoria: {sin_cat}")
    print(f"Sin horario: {sin_hor}")
    wb.close()
    return filas

# -----------------------------------------------------------
# 3) Cruce de personal: num_emp comunes entre ambas fuentes
# -----------------------------------------------------------
def cruzar(p_vac, p_gen):
    dump_section("Cruce PERSONAL (vac vs general)")
    vac_by_num = {f["num_emp"]: f for f in p_vac if f["num_emp"] is not None}
    gen_by_num = {f["num_emp"]: f for f in p_gen if f["num_emp"] is not None}

    solo_vac = set(vac_by_num) - set(gen_by_num)
    solo_gen = set(gen_by_num) - set(vac_by_num)
    en_ambas = set(vac_by_num) & set(gen_by_num)
    print(f"En ambas: {len(en_ambas)}  Solo VAC: {len(solo_vac)}  Solo GEN: {len(solo_gen)}")
    if solo_vac: print(f"  solo VAC: {sorted(solo_vac)}")
    if solo_gen: print(f"  solo GEN: {sorted(solo_gen)}")

    # Comparar nombre normalizado cuando hay match
    discrepancias = []
    for n in en_ambas:
        a = norm_name(vac_by_num[n]["nombre"])
        b = norm_name(gen_by_num[n]["nombre"])
        if a != b:
            discrepancias.append((n, a, b))
    if discrepancias:
        print(f"Discrepancias en nombre (mismo num_emp): {len(discrepancias)}")
        for n,a,b in discrepancias[:10]:
            print(f"  {n}: VAC={a!r}  GEN={b!r}")

# -----------------------------------------------------------
# 4) VACACIONES
# -----------------------------------------------------------
def analizar_vacaciones():
    dump_section("VACACIONES - analisis de las 3 hojas")

    total = 0
    for path, sheet in [
        ("BASE GENERAL.xlsx",          "VACACIONES"),
        ("BASE VACACIONES 2026.xlsx",  "VACACIONES 2025"),
        ("BASE VACACIONES 2026.xlsx",  "VACACIONES SINDICALIZADOS 2026"),
    ]:
        full = os.path.join(ROOT, path)
        wb = load_workbook(full, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            print(f"  [{path} :: {sheet}] no existe")
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else ()
        data = [r for r in rows[1:] if not all(v is None for v in r)]
        print(f"\n[{path} :: {sheet}] filas={len(data)}  cols={len(header)}")

        # estatus implicito: usamos DIAS PENDIENTES y fechas para derivar
        sin_fecha = 0
        sin_nombre = 0
        sin_num = 0
        fechas_raras = 0
        periodos = Counter()
        for r in data:
            # layout es ligeramente distinto en 2026 sindicalizados pero
            # las columnas clave coinciden en posiciones relativas:
            # (CONS, FECHA_INCID, NOMBRE|NUM, NUM|NOMBRE, NOMINA|S, ...)
            # dejamos analisis generico
            if r[1] is None: sin_fecha += 1
            if isinstance(r[1], str):
                try:
                    datetime.strptime(r[1], "%d/%m/%Y")
                except Exception:
                    fechas_raras += 1
            # consolidado
            # Para 2026 sind la col 2 es NO_EMP y col 3 es NOMBRE; en otras es al reves
            peri_col = 7 if sheet == "VACACIONES SINDICALIZADOS 2026" else 6
            periodos[str(r[peri_col]).strip().upper() if r[peri_col] else "?"] += 1
        print(f"  sin fecha: {sin_fecha}")
        print(f"  fechas invalidas (string no parse): {fechas_raras}")
        print(f"  periodos: {dict(periodos)}")
        total += len(data)
        wb.close()
    print(f"\nTotal filas vacaciones consideradas: {total}")

# -----------------------------------------------------------
# 5) FOLIOS INTERNO SUB
# -----------------------------------------------------------
def analizar_folios_internos():
    dump_section("FOLIOS INTERNO SUB - BASE GENERAL.xlsx")
    path = os.path.join(ROOT, "BASE GENERAL.xlsx")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["FOLIOS INTERNO SUB"]
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[1:] if not all(v is None for v in r)]
    print(f"filas={len(data)}")

    # duplicados por (fecha, folio_direccion, asunto)
    keys = []
    status = Counter()
    for r in data:
        f = r[0]
        fd = (r[3] or "").strip().upper() if len(r)>3 else ""
        asu = (r[4] or "").strip().upper() if len(r)>4 else ""
        keys.append((str(f)[:10] if f else "", fd, asu[:40]))
        st = (r[9] or "").strip().upper() if len(r)>9 else ""
        status[st] += 1
    dup = [k for k,c in Counter(keys).items() if c > 1]
    print(f"duplicados por (fecha, folio_direccion, asunto_prefix): {len(dup)}")
    for d in dup[:5]: print(f"  {d}")

    print(f"status: {dict(status)}")

    # nombres de direccion unicos (para resolver areas_internas)
    dirs = Counter()
    for r in data:
        nd = (r[2] or "").strip().upper() if len(r)>2 else ""
        dirs[nd] += 1
    print(f"direcciones distintas: {len(dirs)}")
    for k,v in dirs.most_common():
        print(f"  {v:>3} x  {k}")
    wb.close()

def main():
    p_vac = analizar_personal_vac()
    p_gen = analizar_personal_gen()
    cruzar(p_vac, p_gen)
    analizar_vacaciones()
    analizar_folios_internos()

if __name__ == "__main__":
    main()
