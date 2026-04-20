"""
Inspeccion focalizada en columnas E/S/DC/INC de la hoja INCIDENCIAS.
Detecta combinaciones de marcas, duplicados, filas sin tipo.
"""
from __future__ import annotations
import os
import unicodedata
from collections import Counter
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def sa(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()


def norm_mark(v):
    if v is None:
        return ""
    s = str(v).strip().upper()
    return s


path = os.path.join(ROOT, "BASE GENERAL.xlsx")
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb["INCIDENCIAS"]
rows = list(ws.iter_rows(values_only=True))
data = [r for r in rows[1:] if not all(v is None for v in r)]
print(f"Filas con datos: {len(data)}")

# Filtrar solo filas con num_empleado Y nombre
filas_reales = []
filas_vacias = 0
for r in data:
    num = r[2] if len(r) > 2 else None
    nombre = r[3] if len(r) > 3 else None
    if (num is None or str(num).strip() == "") and (nombre is None or str(nombre).strip() == ""):
        filas_vacias += 1
        continue
    filas_reales.append(r)
print(f"Filas con empleado identificable: {len(filas_reales)}")
print(f"Filas vacias (sin num ni nombre): {filas_vacias}")

# Combinaciones E/S/DC/INC
combos = Counter()
sin_marca = 0
multi_marca = 0
for r in filas_reales:
    E = norm_mark(r[6] if len(r) > 6 else None)
    S = norm_mark(r[7] if len(r) > 7 else None)
    DC = norm_mark(r[8] if len(r) > 8 else None)
    INC = norm_mark(r[9] if len(r) > 9 else None)
    marks = []
    if E: marks.append("E")
    if S: marks.append("S")
    if DC: marks.append("DC")
    if INC: marks.append("INC")
    if not marks:
        sin_marca += 1
    if len(marks) > 1:
        multi_marca += 1
    key = "+".join(marks) if marks else "SIN_MARCA"
    combos[key] += 1
print(f"\nCombinaciones E/S/DC/INC: {dict(combos)}")
print(f"Filas sin ninguna marca: {sin_marca}")
print(f"Filas con multiples marcas: {multi_marca}")

# Valores raros en columnas de marca
print("\nValores distintos en columnas de marca:")
for idx, nombre_col in [(6, "E"), (7, "S"), (8, "DC"), (9, "INC")]:
    vals = Counter()
    for r in filas_reales:
        v = r[idx] if len(r) > idx else None
        vals[norm_mark(v)] += 1
    print(f"  col {nombre_col:<4} -> {dict(vals)}")

# Tipo personal (S/C)
print("\nDistribucion S/C (col 4):")
tp = Counter()
for r in filas_reales:
    v = r[4] if len(r) > 4 else None
    tp[norm_mark(v)] += 1
print(f"  {dict(tp)}")

# num_empleado duplicados (NO es duplicado por si solo, pero util para ver)
nums = Counter()
for r in filas_reales:
    v = r[2] if len(r) > 2 else None
    if v is not None:
        nums[str(v).strip()] += 1
print(f"\nTotal num_empleado distintos: {len(nums)}")
print(f"Empleado con mas incidencias (top 10):")
for k, c in nums.most_common(10):
    print(f"  {c:>3}  {k}")

# Fechas
fechas_incid = Counter()
fechas_just = Counter()
fechas_recib = Counter()
for r in filas_reales:
    for idx, target in [(1, fechas_incid), (5, fechas_just), (11, fechas_recib)]:
        v = r[idx] if len(r) > idx else None
        if v is None:
            target["<NULL>"] += 1
        elif isinstance(v, (datetime, date)):
            target["<FECHA>"] += 1
        else:
            target[f"<TEXTO:{type(v).__name__}>"] += 1

print(f"\nfecha_incidencia: {dict(fechas_incid)}")
print(f"dia_justificar:    {dict(fechas_just)}")
print(f"recibido_coord:    {dict(fechas_recib)}")

# Duplicados por (num, fecha_incid, E/S/DC/INC)
firmas = Counter()
for r in filas_reales:
    num = str(r[2]).strip() if r[2] is not None else ""
    fi = r[1]
    fi_s = fi.strftime("%Y-%m-%d") if isinstance(fi, (datetime, date)) else str(fi)
    fj = r[5]
    fj_s = fj.strftime("%Y-%m-%d") if isinstance(fj, (datetime, date)) else str(fj)
    tipo_bits = (norm_mark(r[6]), norm_mark(r[7]), norm_mark(r[8]), norm_mark(r[9]))
    firmas[(num, fi_s, fj_s, tipo_bits)] += 1
dups = [(k, c) for k, c in firmas.items() if c > 1]
print(f"\nDuplicados por (num_emp, fecha_incid, dia_just, tipo_bits): {len(dups)}")
for k, c in dups[:5]:
    print(f"  {c}x  {k}")

# Muestra con multi-marcas
print("\nFilas con multiples marcas (primeras 5):")
n = 0
for i, r in enumerate(filas_reales, start=1):
    if n >= 5:
        break
    E = norm_mark(r[6] if len(r) > 6 else None)
    S = norm_mark(r[7] if len(r) > 7 else None)
    DC = norm_mark(r[8] if len(r) > 8 else None)
    INC = norm_mark(r[9] if len(r) > 9 else None)
    marks = [m for m, v in [("E", E), ("S", S), ("DC", DC), ("INC", INC)] if v]
    if len(marks) > 1:
        print(f"  fila_real {i} marcas={marks} : {r}")
        n += 1

# Muestra sin marca
print("\nFilas sin marca alguna (primeras 5):")
n = 0
for i, r in enumerate(filas_reales, start=1):
    if n >= 5:
        break
    E = norm_mark(r[6] if len(r) > 6 else None)
    S = norm_mark(r[7] if len(r) > 7 else None)
    DC = norm_mark(r[8] if len(r) > 8 else None)
    INC = norm_mark(r[9] if len(r) > 9 else None)
    marks = [m for m, v in [("E", E), ("S", S), ("DC", DC), ("INC", INC)] if v]
    if not marks:
        print(f"  fila_real {i}: {r}")
        n += 1

wb.close()
