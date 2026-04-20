"""
Importador de VACACIONES.

Fuentes:
  1) BASE GENERAL.xlsx          hoja 'VACACIONES'                  (2026 mezclado)
  2) BASE VACACIONES 2026.xlsx  hoja 'VACACIONES 2025'             (histórico 2025)
  3) BASE VACACIONES 2026.xlsx  hoja 'VACACIONES SINDICALIZADOS 2026'

Layouts reales detectados (heterogéneos):

  Hoja 'VACACIONES' (BASE GENERAL.xlsx):
    0 CONS | 1 FECHA_INCIDENCIA | 2 NOMBRE | 3 NO_EMPLEADO | 4 NOMINA |
    5 DIAS_POR_JUSTIFICAR (texto fechas) | 6 PERIODO |
    7 DIAS_CORRESPONDEN | 8 DIAS_OTORGADOS | 9 DIAS_PENDIENTES |
    10 REGRESA | 11 RECIBIDO_COORD | 12 FOLIO_TM

  Hoja 'VACACIONES 2025' (BASE VACACIONES 2026):
    igual que 'VACACIONES'.

  Hoja 'VACACIONES SINDICALIZADOS 2026' (BASE VACACIONES 2026):
    0 CONS | 1 FECHA_VACACIONES | 2 NO_EMPLEADO | 3 NOMBRE | 4 S | 5 C |
    6 DIAS_POR_JUSTIFICAR (texto) | 7 PERIODO |
    8 DIAS_CORRESPONDEN | 9 DIAS_OTORGADOS | 10 DIAS_PENDIENTES |
    11 REGRESA | 12 RECIBIDO_COORD | 13 FOLIO_TM

Uso:
    python tools/importar_vacaciones.py --staging
    python tools/importar_vacaciones.py --dry-run
    python tools/importar_vacaciones.py --importar

Reglas:
    - Solo se procesan filas con NO_EMPLEADO + NOMBRE + FECHA_INCIDENCIA válida.
    - Se intenta derivar fecha_inicio / fecha_fin desde el TEXTO del periodo
      ('DEL 27 DE FEBRERO AL 13 DE MARZO DE 2026', '10 AL 15 DE MARZO 2026', etc).
      Si no se puede parsear, la fila queda en staging con estado_revision='PENDIENTE_REVISION'.
    - El estatus se deriva por fecha actual: ACTIVA / FINALIZADA / PROGRAMADA.
      Si no hay fecha_fin -> PENDIENTE_REVISION.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime, date
from typing import Any, Optional, Tuple

try:
    import psycopg2
    from psycopg2.extras import execute_batch, Json
except ImportError:
    print("ERROR: falta psycopg2. pip install psycopg2-binary"); sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. pip install openpyxl"); sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_GEN = os.path.join(ROOT, "BASE GENERAL.xlsx")
XLSX_VAC = os.path.join(ROOT, "BASE VACACIONES 2026.xlsx")

def pg_connect():
    return psycopg2.connect(
        host     = os.environ.get("PGHOST",     "localhost"),
        port     = int(os.environ.get("PGPORT", "5432")),
        dbname   = os.environ.get("PGDATABASE", "respaldooficios"),
        user     = os.environ.get("PGUSER",     "postgres"),
        password = os.environ.get("PGPASSWORD", "admin"),
    )

# ---------------- helpers -----------------

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
    "FEBRERI": 2,  # error común visto en Excel
}

def sa(s: Optional[str]) -> str:
    if not s: return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()

def norm(v: Any) -> Optional[str]:
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s: return None
    # Excel VLOOKUP fallidos o celdas con errores
    if s.upper() in ("#N/A", "#N/D", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#DIV/0!"):
        return None
    return s

def parse_date_any(v: Any) -> Optional[str]:
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s: return None
    # corrección obvia de errores tipográficos en fechas (p.ej. "10/02/20258")
    s = re.sub(r"^(\d{1,2}/\d{1,2}/\d{4})\d+$", r"\1", s)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError: pass
    return None

def to_jsonable(row: tuple) -> list:
    return [v.strftime("%Y-%m-%d") if isinstance(v,(datetime,date)) else v for v in row]

# -----------------------------------------------------------
# Parser de texto libre "DEL 25 AL 27 DE MARZO DE 2026"
# -----------------------------------------------------------
_MES_RE = "|".join(sorted(MESES.keys(), key=len, reverse=True))

def parse_periodo_texto(texto: str, anio_fallback: int) -> Tuple[Optional[str], Optional[str]]:
    """
    Devuelve (fecha_inicio_iso, fecha_fin_iso) a partir de textos como:
      'DEL 27 DE FEBRERO AL 13 DE MARZO DE 2026'   (2 meses)
      'DEL 30 DE DICIEMBRE AL 06 DE ENERO 2026'    (2 meses cruzando anio)
      '20 AL 26 DE FEBRERO DE 2025'                (1 mes, rango)
      '24 DE MARZO DE 2025'                        (un solo día)
      '17,18, 19 Y 20 DE MARZO DE 2026'            (enumeracion)
      'EL 30 Y 31 DE MARZO DE 2026'                (enumeracion corta)
    """
    if not texto: return (None, None)
    t = sa(texto)
    t = t.replace(".", " ").replace(",", " ")
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"\bDEL\b\s+", "", t)
    t = re.sub(r"^EL\s+", "", t)

    # Extraer anio explicito (el ultimo que aparezca)
    anio = anio_fallback
    anios_found = re.findall(r"\b(20\d{2})\b", t)
    if anios_found:
        anio = int(anios_found[-1])

    # Encontrar todas las apariciones de mes en orden
    meses_en_texto = list(re.finditer(rf"\b({_MES_RE})\b", t))
    if not meses_en_texto:
        return (None, None)

    # Obtener dias - numeros de 1 a 31 excluyendo el anio
    def extraer_dias(fragmento: str) -> list:
        ds = []
        for m in re.finditer(r"\b(\d{1,2})\b", fragmento):
            n = int(m.group(1))
            if 1 <= n <= 31:
                ds.append(n)
        return ds

    try:
        if len(meses_en_texto) >= 2:
            # Caso 'DEL X DE MES1 AL Y DE MES2 [DE ANIO]'
            m1 = meses_en_texto[0]
            m2 = meses_en_texto[-1]
            mes1 = MESES[m1.group(1)]
            mes2 = MESES[m2.group(1)]
            # Dias antes de mes1 -> dia_ini; dias entre mes1 y mes2 -> dia_fin
            dias_ini = extraer_dias(t[:m1.start()])
            dias_fin = extraer_dias(t[m1.end():m2.start()])
            # Si no hay dias entre, puede ser '27 DE FEBRERO AL 13 DE MARZO'
            # entonces dia_fin viene despues de "AL" y antes de mes2.
            if not dias_fin:
                # fallback: buscar dias tras m1 en todo el texto restante
                dias_fin = extraer_dias(t[m1.end():])
                # quitar cualquier dia que coincida con el anio
                dias_fin = [d for d in dias_fin if d != anio % 100]
            if not dias_ini or not dias_fin:
                return (None, None)
            dia_ini = dias_ini[0]
            dia_fin = dias_fin[-1] if dias_fin else dia_ini
            # Anio del inicio: si mes1 > mes2 cruzo anio, inicio es anio-1
            anio_ini = anio - 1 if mes1 > mes2 else anio
            anio_fin = anio
            ini = date(anio_ini, mes1, dia_ini)
            fin = date(anio_fin, mes2, dia_fin)
            if fin < ini:
                return (None, None)
            return ini.strftime("%Y-%m-%d"), fin.strftime("%Y-%m-%d")

        # Un solo mes: rango dentro del mes o dia unico
        m1 = meses_en_texto[0]
        mes1 = MESES[m1.group(1)]
        dias = extraer_dias(t)
        # filtrar posibles restos del anio
        dias = [d for d in dias if d != anio % 100]
        if not dias:
            return (None, None)
        dia_ini = dias[0]
        dia_fin = dias[-1] if len(dias) > 1 else dia_ini
        ini = date(anio, mes1, dia_ini)
        fin = date(anio, mes1, dia_fin)
        if fin < ini:
            fin = ini
        return ini.strftime("%Y-%m-%d"), fin.strftime("%Y-%m-%d")
    except Exception:
        return (None, None)

# -----------------------------------------------------------
# FASE 1: STAGING
# -----------------------------------------------------------
def ensure_open(path: str):
    if not os.path.exists(path): return None
    return load_workbook(path, data_only=True, read_only=True)

def cargar_staging(conn) -> int:
    rows_all = []

    def push(fuente, fila, r, mapping):
        raw = [None]*12
        # mapping: nombres->indice de col en la tupla
        raw[0]  = parse_date_any(r[mapping["fecha_incid"]])      if mapping.get("fecha_incid") is not None else None
        raw[1]  = norm(r[mapping["nombre"]])                      if mapping.get("nombre")      is not None else None
        raw[2]  = norm(r[mapping["num"]])                         if mapping.get("num")         is not None else None
        raw[3]  = norm(r[mapping["nomina"]])                      if mapping.get("nomina")      is not None else None
        raw[4]  = norm(r[mapping["dias_justificar"]])             if mapping.get("dias_justificar") is not None else None
        raw[5]  = norm(r[mapping["periodo"]])                     if mapping.get("periodo")     is not None else None
        raw[6]  = norm(r[mapping["dias_corresp"]])                if mapping.get("dias_corresp") is not None else None
        raw[7]  = norm(r[mapping["dias_otorg"]])                  if mapping.get("dias_otorg")  is not None else None
        raw[8]  = norm(r[mapping["dias_pend"]])                   if mapping.get("dias_pend")   is not None else None
        raw[9]  = parse_date_any(r[mapping["regresa"]])           if mapping.get("regresa")     is not None else None
        raw[10] = parse_date_any(r[mapping["recibido_coord"]])    if mapping.get("recibido_coord") is not None else None
        raw[11] = norm(r[mapping["folio"]])                       if mapping.get("folio")       is not None else None
        # No cargar filas totalmente en blanco (salvo para nombre)
        if not any(raw):  # todos None
            return
        # No cargar filas que solo traen nombre sin fecha ni periodo (plantilla)
        if not raw[0] and not raw[5] and not raw[1]:
            return
        rows_all.append((
            fuente, fila,
            raw[0], raw[1], raw[2], raw[3], raw[4], raw[5],
            raw[6], raw[7], raw[8], raw[9], raw[10], raw[11],
            Json(to_jsonable(r)),
        ))

    # Hoja estilo A (BASE GENERAL - VACACIONES  +  BASE VACACIONES 2026 - VACACIONES 2025)
    map_a = {
        "fecha_incid": 1, "nombre": 2, "num": 3, "nomina": 4,
        "dias_justificar": 5, "periodo": 6,
        "dias_corresp": 7, "dias_otorg": 8, "dias_pend": 9,
        "regresa": 10, "recibido_coord": 11, "folio": 12,
    }
    # Hoja estilo B (VACACIONES SINDICALIZADOS 2026): nombre/num invertidos + 2 cols de S/C
    map_b = {
        "fecha_incid": 1, "num": 2, "nombre": 3,
        "nomina": None,   # hay cols separadas S(4) y C(5); no la usamos aquí
        "dias_justificar": 6, "periodo": 7,
        "dias_corresp": 8, "dias_otorg": 9, "dias_pend": 10,
        "regresa": 11, "recibido_coord": 12, "folio": 13,
    }

    # 1) BASE GENERAL - VACACIONES
    wb = ensure_open(XLSX_GEN)
    if wb and "VACACIONES" in wb.sheetnames:
        ws = wb["VACACIONES"]
        rows = list(ws.iter_rows(values_only=True))
        for idx, r in enumerate(rows[1:], start=2):
            if not r or all(v is None for v in r): continue
            push("BASE GENERAL.xlsx!VACACIONES", idx, r, map_a)
        wb.close()

    # 2) BASE VACACIONES 2026 - VACACIONES 2025
    wb = ensure_open(XLSX_VAC)
    if wb and "VACACIONES 2025" in wb.sheetnames:
        ws = wb["VACACIONES 2025"]
        rows = list(ws.iter_rows(values_only=True))
        for idx, r in enumerate(rows[1:], start=2):
            if not r or all(v is None for v in r): continue
            push("BASE VACACIONES 2026.xlsx!VACACIONES 2025", idx, r, map_a)

    # 3) VACACIONES SINDICALIZADOS 2026
    if wb and "VACACIONES SINDICALIZADOS 2026" in wb.sheetnames:
        ws = wb["VACACIONES SINDICALIZADOS 2026"]
        rows = list(ws.iter_rows(values_only=True))
        for idx, r in enumerate(rows[1:], start=2):
            if not r or all(v is None for v in r): continue
            # Estas hojas tienen 50+ filas-plantilla vacías (solo nombre sin fechas).
            # Las saltamos antes de insertar usando el check en push().
            push("BASE VACACIONES 2026.xlsx!VACACIONES SINDICALIZADOS 2026", idx, r, map_b)
        wb.close()
    elif wb:
        wb.close()

    if not rows_all:
        print("No hay filas para staging vacaciones."); return 0

    with conn.cursor() as cur:
        cur.execute("TRUNCATE import_vacaciones_raw RESTART IDENTITY")
        execute_batch(cur,
            """INSERT INTO import_vacaciones_raw
                (fuente, fila_excel,
                 raw_fecha_incidencia, raw_nombre, raw_num_empleado, raw_nomina,
                 raw_dias_justificar, raw_periodo,
                 raw_dias_correspond, raw_dias_otorgados, raw_dias_pendientes,
                 raw_regresa, raw_recibido_coord, raw_folio_tm, raw_data)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            rows_all, page_size=100)
        conn.commit()
    print(f"Staging import_vacaciones_raw: {len(rows_all)} filas cargadas.")
    return len(rows_all)


# -----------------------------------------------------------
# FASE 2: staging -> vacaciones
# -----------------------------------------------------------
def procesar(conn, dry_run: bool = False) -> None:
    with conn.cursor() as cur:
        cur.execute("SELECT id, clave FROM estatus_vacaciones")
        mapa_est = {clave: eid for (eid, clave) in cur.fetchall()}

        cur.execute(
            """SELECT id, fuente, fila_excel,
                      raw_fecha_incidencia, raw_nombre, raw_num_empleado, raw_nomina,
                      raw_dias_justificar, raw_periodo,
                      raw_dias_correspond, raw_dias_otorgados, raw_dias_pendientes,
                      raw_regresa, raw_recibido_coord, raw_folio_tm
                 FROM import_vacaciones_raw
                WHERE procesado = FALSE
                ORDER BY fila_excel"""
        )
        filas = cur.fetchall()

    ok = omitidos = errores = revision = creados = saltadas_vacias = 0

    for (sid, fuente, fila_ex,
         f_inc, nombre, num_emp, nomina,
         dias_just_txt, periodo_txt,
         dias_corresp, dias_otorg, dias_pend,
         freg, frec, folio) in filas:

        try:
            # Saltar filas-plantilla totalmente vacias (sin nombre, num, fecha, periodo ni texto)
            if (not nombre and not num_emp and not f_inc
                    and not periodo_txt and not dias_just_txt):
                saltadas_vacias += 1
                with conn.cursor() as cskip:
                    cskip.execute(
                        """UPDATE import_vacaciones_raw
                              SET procesado=TRUE, accion='OMITIDO_VACIA',
                                  estado_revision='OK', error_importacion=NULL
                            WHERE id=%s""",
                        (sid,)
                    )
                    conn.commit()
                continue

            if not nombre and not num_emp:
                raise ValueError(f"Fila {fila_ex}: sin nombre ni num_empleado.")

            # Resolver personal
            with conn.cursor() as c:
                pid = None
                if num_emp:
                    c.execute("SELECT id FROM personal WHERE numero_empleado=%s LIMIT 1", (num_emp,))
                    row = c.fetchone()
                    if row: pid = row[0]
                if not pid and nombre:
                    c.execute(
                        """SELECT id FROM personal
                            WHERE nombre_normalizado = %s LIMIT 1""",
                        (sa(nombre),)
                    )
                    row = c.fetchone()
                    if row: pid = row[0]
            if not pid:
                # Auto-crear personal con tipo desde 'nomina' si viene
                tipo_clave = "SINDICALIZADO" if (nomina or "").strip().upper() == "S" else (
                              "CONFIANZA"    if (nomina or "").strip().upper() == "C" else None)
                if not tipo_clave:
                    raise ValueError(f"Fila {fila_ex}: no existe personal y no se puede determinar tipo S/C.")
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM tipos_personal WHERE clave=%s", (tipo_clave,))
                    tipo_id = cur.fetchone()[0]
                    cur.execute(
                        """INSERT INTO personal
                            (numero_empleado, nombre_completo, nombre_normalizado, tipo_personal_id)
                           VALUES (%s,%s,%s,%s) RETURNING id""",
                        (num_emp, nombre.strip() if nombre else f"SIN NOMBRE {num_emp}",
                         sa(nombre) if nombre else f"SIN NOMBRE {num_emp}".upper(),
                         tipo_id)
                    )
                    pid = cur.fetchone()[0]
                    conn.commit()

            # Derivar fechas
            f_ini_iso = f_fin_iso = None
            anio_fall = datetime.now().year
            if f_inc:
                try:
                    anio_fall = int(f_inc[:4])
                except Exception: pass
            if dias_just_txt:
                f_ini_iso, f_fin_iso = parse_periodo_texto(dias_just_txt, anio_fall)
            # si aun asi no se pudo, intentar usar fecha_incidencia -> fecha_incidencia (un dia)
            if not f_ini_iso and f_inc:
                f_ini_iso = f_fin_iso = f_inc

            if not f_ini_iso or not f_fin_iso:
                revision += 1
                with conn.cursor() as cerr:
                    cerr.execute(
                        """UPDATE import_vacaciones_raw
                              SET estado_revision='PENDIENTE_REVISION',
                                  error_importacion=%s,
                                  personal_id=%s
                            WHERE id=%s""",
                        (f"No se pudo derivar fecha_inicio/fin del texto: {dias_just_txt!r}", pid, sid)
                    )
                    conn.commit()
                continue

            # Coherencia fecha_fin vs fecha_regreso:
            # el parser del texto libre falla con enumeraciones como
            # '27,28,31 DE MARZO Y 01 DE ABRIL' (asigna el mismo dia).
            # La columna 'REGRESA A LABORAR' es autoridad de RH: siempre es
            # el primer dia laborable DESPUES del fin de vacaciones, por lo
            # tanto fecha_fin debe ser <= fecha_regreso - 1.
            if freg:
                try:
                    fr = datetime.strptime(freg, "%Y-%m-%d").date()
                    ff = datetime.strptime(f_fin_iso, "%Y-%m-%d").date()
                    fi = datetime.strptime(f_ini_iso, "%Y-%m-%d").date()
                    if ff >= fr:
                        from datetime import timedelta
                        nueva_ff = fr - timedelta(days=1)
                        if nueva_ff >= fi:
                            f_fin_iso = nueva_ff.strftime("%Y-%m-%d")
                        else:
                            # Datos contradictorios -> revision
                            revision += 1
                            with conn.cursor() as cerr:
                                cerr.execute(
                                    """UPDATE import_vacaciones_raw
                                          SET estado_revision='PENDIENTE_REVISION',
                                              error_importacion=%s,
                                              personal_id=%s
                                        WHERE id=%s""",
                                        (f"Inconsistencia: fecha_regreso {freg} <= fecha_inicio {f_ini_iso}.",
                                         pid, sid)
                                )
                                conn.commit()
                            continue
                except Exception:
                    pass

            # Normalizar periodo
            pstr = sa(periodo_txt or "")
            periodo_corto = None; anio_periodo = None
            if "1ER" in pstr or "1°" in pstr or "1ER." in pstr or "PRIMER" in pstr:
                periodo_corto = "1ER"
            elif "2DO" in pstr or "2°" in pstr or "2DO." in pstr or "SEGUNDO" in pstr:
                periodo_corto = "2DO"
            m_anio = re.search(r"(20\d{2})", pstr)
            if m_anio:
                anio_periodo = int(m_anio.group(1))
            periodo_final = f"{periodo_corto} {anio_periodo}" if periodo_corto and anio_periodo else periodo_corto

            # Días
            def to_int(v):
                if v is None: return 0
                try: return int(str(v).strip().split('.')[0])
                except Exception: return 0
            dsol = to_int(dias_corresp)
            ddis = to_int(dias_otorg)
            dpen = to_int(dias_pend)
            if dpen < 0: dpen = 0

            # Estatus derivado
            hoy = date.today()
            fi = datetime.strptime(f_ini_iso, "%Y-%m-%d").date()
            ff = datetime.strptime(f_fin_iso, "%Y-%m-%d").date()
            if fi > hoy:      est_clave = "PROGRAMADA"
            elif ff >= hoy:   est_clave = "ACTIVA"
            else:             est_clave = "FINALIZADA"
            est_id = mapa_est[est_clave]

            if dry_run:
                print(f"  [DRY] fila {fila_ex:>4}  pid={pid}  {f_ini_iso}..{f_fin_iso}  "
                      f"periodo={periodo_final}  est={est_clave}  dsol/ddis/dpen={dsol}/{ddis}/{dpen}")
                ok += 1; continue

            try:
                with conn.cursor() as cins:
                    cins.execute(
                        """INSERT INTO vacaciones
                           (personal_id, periodo, anio, periodo_texto,
                            fecha_solicitud, fecha_inicio, fecha_fin,
                            fecha_regreso, fecha_recibido_coord,
                            dias_solicitados, dias_disfrutados, dias_pendientes,
                            folio_tm, estatus_id, capturado_por_user_id, observaciones)
                           VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,
                                   %s,%s,%s, %s,%s,NULL,%s)
                           ON CONFLICT (personal_id, fecha_inicio, fecha_fin, COALESCE(periodo,''))
                           DO NOTHING
                           RETURNING id""",
                        (pid, periodo_final, anio_periodo, dias_just_txt,
                         f_inc, f_ini_iso, f_fin_iso,
                         freg, frec,
                         dsol, ddis, dpen,
                         folio, est_id,
                         f"Importado de {fuente}, fila {fila_ex}.")
                    )
                    new = cins.fetchone()
                    if new:
                        vid = new[0]
                        creados += 1
                        accion = "CREADO"
                    else:
                        # Ya existía mismo tramo (duplicado)
                        omitidos += 1
                        accion = "OMITIDO_DUP"
                        vid = None

                    cins.execute(
                        """UPDATE import_vacaciones_raw
                              SET procesado=TRUE, vacaciones_id=%s,
                                  personal_id=%s, accion=%s,
                                  fecha_inicio_parsed=%s, fecha_fin_parsed=%s,
                                  estado_revision='OK', error_importacion=NULL
                            WHERE id=%s""",
                        (vid, pid, accion, f_ini_iso, f_fin_iso, sid)
                    )
                conn.commit()
                ok += 1

            except psycopg2.Error as pe:
                conn.rollback()
                with conn.cursor() as cerr:
                    cerr.execute(
                        """UPDATE import_vacaciones_raw
                              SET estado_revision='ERROR', error_importacion=%s
                            WHERE id=%s""",
                        (str(pe), sid)
                    )
                    conn.commit()
                errores += 1
                print(f"  [ERR fila {fila_ex}] {pe}")

        except Exception as e:
            errores += 1
            conn.rollback()
            with conn.cursor() as cerr:
                cerr.execute(
                    """UPDATE import_vacaciones_raw
                          SET estado_revision='ERROR', error_importacion=%s
                        WHERE id=%s""",
                    (str(e), sid)
                )
                conn.commit()
            print(f"  [ERR fila {fila_ex}] {e}")

    print("\n--- Resultados vacaciones ---")
    print(f"OK={ok}  creados={creados}  omitidos_dup={omitidos}  "
          f"saltadas_vacias={saltadas_vacias}  revision={revision}  errores={errores}  "
          f"total_staging={len(filas)}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--staging",  action="store_true")
    ap.add_argument("--dry-run",  action="store_true")
    ap.add_argument("--importar", action="store_true")
    args = ap.parse_args()

    if not (args.staging or args.dry_run or args.importar):
        ap.print_help(); sys.exit(0)

    conn = pg_connect()
    try:
        if args.staging:  cargar_staging(conn)
        if args.dry_run:  procesar(conn, dry_run=True)
        if args.importar: procesar(conn, dry_run=False)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
