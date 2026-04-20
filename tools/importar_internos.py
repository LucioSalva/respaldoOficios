"""
Importador de la hoja FOLIOS INTERNO SUB del Excel BASE GENERAL.xlsx.

Uso:
    # 1) Llenar SOLO la staging (seguro, NO toca oficios finales):
    python tools/importar_internos.py --staging

    # 2) Ver un dry-run de la normalización (no escribe nada en oficios):
    python tools/importar_internos.py --dry-run

    # 3) Importar a oficios (INTERNO) leyendo desde la staging:
    python tools/importar_internos.py --importar

Variables de entorno (o valores por defecto para ambiente local):
    PGHOST=localhost PGPORT=5432 PGUSER=postgres PGPASSWORD=admin
    PGDATABASE=respaldooficios
"""
from __future__ import annotations

import argparse
import os
import sys
import unicodedata
from datetime import datetime, date
from typing import Any, Optional

try:
    import psycopg2
    from psycopg2.extras import execute_batch
except ImportError:
    print("ERROR: falta psycopg2. Instalar con:  pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. Instalar con:  pip install openpyxl")
    sys.exit(1)

ROOT   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX   = os.path.join(ROOT, "BASE GENERAL.xlsx")
SHEET  = "FOLIOS INTERNO SUB"

# Usuario "lucio" se asume como usuario_capturo_id para los registros importados.
USUARIO_CAPTURO = "lucio"

def pg_connect():
    return psycopg2.connect(
        host     = os.environ.get("PGHOST",     "localhost"),
        port     = int(os.environ.get("PGPORT", "5432")),
        dbname   = os.environ.get("PGDATABASE", "respaldooficios"),
        user     = os.environ.get("PGUSER",     "postgres"),
        password = os.environ.get("PGPASSWORD", "admin"),
    )

# --------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------
def norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None

def sin_acentos(s: str) -> str:
    if s is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch)).upper().strip()

def is_empty_row(row) -> bool:
    return all(norm(c) is None for c in row)

def parse_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    # Intentar varios formatos comunes
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

# Mapeo manual de variantes de "NOMBRE DIRECCIÓN" a nombre canónico en areas_internas.
AREA_ALIAS = {
    "TESORERIA MUNICIPAL":                                      "TESORERÍA MUNICIPAL",
    "TESORERO MUNICIPAL":                                       "TESORERÍA MUNICIPAL",
    "COORDINACION ADMINISTRATIVA DE TESORERIA":                 "COORDINACIÓN ADMINISTRATIVA DE TESORERÍA",
    "SUBDIRECCION DE EGRESOS":                                  "SUBDIRECCIÓN DE EGRESOS",
    "SUBDIRECCION DE JURIDICO DE LA TESORERIA":                 "SUBDIRECCIÓN DE JURÍDICO DE TESORERÍA",
    "SUBDIRECCION DE PROGRAMAS FEDERALES":                      "SUBDIRECCIÓN DE PROGRAMAS FEDERALES Y ESTATALES",
    "SUBDIRECCION DE PROGRAMAS FEDERALES Y ESTATLES":           "SUBDIRECCIÓN DE PROGRAMAS FEDERALES Y ESTATALES",
    "SUBDIRECCION DE PROGRAMAS FEDERALES Y ESTATALES":          "SUBDIRECCIÓN DE PROGRAMAS FEDERALES Y ESTATALES",
    "CORDINADOR DE LAS OFICIALIAS DE REGISTRO CIVIL DEL H. AYUNTAMIENTO": "COORDINACIÓN DE OFICIALÍAS DE REGISTRO CIVIL",
    "COORDINACION DE OFICIALIAS DE REGISTRO CIVIL":             "COORDINACIÓN DE OFICIALÍAS DE REGISTRO CIVIL",
    "DEPARTAMENTO DE CONTROL DE CAJAS Y RECEPTORIAS EXTERNAS":  "DEPARTAMENTO DE CONTROL DE CAJAS Y RECEPTORÍAS EXTERNAS",
}

def resolver_area(nombre_raw: Optional[str], cache: dict, cur) -> Optional[int]:
    """
    Devuelve id de areas_internas resolviendo acentos/variantes.
    Si no existe, la crea (seed dinámico).
    """
    if not nombre_raw:
        return None
    clave_busq = sin_acentos(nombre_raw)
    alias = AREA_ALIAS.get(clave_busq, nombre_raw.strip())
    # Buscar ignorando acentos y mayúsculas
    if clave_busq in cache:
        return cache[clave_busq]
    cur.execute(
        "SELECT id FROM areas_internas "
        " WHERE upper(unaccent(nombre)) = upper(unaccent(%s)) "
        "    OR upper(unaccent(nombre)) = upper(unaccent(%s))"
        " LIMIT 1",
        (alias, nombre_raw)
    )
    # Si no hay extensión unaccent, caemos al fallback
    row = None
    try:
        row = cur.fetchone()
    except Exception:
        row = None
    if not row:
        # fallback sin unaccent
        cur.execute(
            "SELECT id, nombre FROM areas_internas"
        )
        for (aid, anombre) in cur.fetchall():
            if sin_acentos(anombre) == clave_busq or sin_acentos(anombre) == sin_acentos(alias):
                cache[clave_busq] = aid
                return aid
        # No existe: crear con el nombre alias (canónico)
        cur.execute(
            "INSERT INTO areas_internas (nombre) VALUES (%s) RETURNING id",
            (alias,)
        )
        aid = cur.fetchone()[0]
        cache[clave_busq] = aid
        print(f"  [area creada] {alias}")
        return aid
    cache[clave_busq] = row[0]
    return row[0]

# --------------------------------------------------------------------
# PASO 1: llenar staging
# --------------------------------------------------------------------
def llenar_staging(conn) -> int:
    print(f"Leyendo Excel: {XLSX}")
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET}' no existe.")
        return 0
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Hoja vacía.")
        return 0

    header = rows[0]
    data_rows = []
    for idx, r in enumerate(rows[1:], start=2):   # idx = fila real del Excel
        if is_empty_row(r):
            continue
        data_rows.append((
            idx,                                                  # fila_excel
            parse_date(r[0] if len(r) > 0 else None),             # fecha_recibido
            norm(r[1] if len(r) > 1 else None),                   # folio_minutario
            norm(r[2] if len(r) > 2 else None),                   # nombre_direccion
            norm(r[3] if len(r) > 3 else None),                   # folio_direccion
            norm(r[4] if len(r) > 4 else None),                   # asunto
            parse_date(r[5] if len(r) > 5 else None),             # fecha_oficio_tics
            norm(r[6] if len(r) > 6 else None),                   # folio_oficio_tics
            norm(r[7] if len(r) > 7 else None),                   # realizo
            parse_date(r[8] if len(r) > 8 else None),             # acuse_recibido
            norm(r[9] if len(r) > 9 else None),                   # status
            norm(r[10] if len(r) > 10 else None),                 # capturo
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE import_folios_interno_sub_raw RESTART IDENTITY CASCADE")
        execute_batch(cur,
            """INSERT INTO import_folios_interno_sub_raw
               (fila_excel, fecha_recibido, folio_minutario, nombre_direccion,
                folio_direccion, asunto, fecha_oficio_tics, folio_oficio_tics,
                realizo, acuse_recibido, status, capturo)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            data_rows, page_size=100)
        conn.commit()
    print(f"Staging cargada: {len(data_rows)} filas.")
    return len(data_rows)

# --------------------------------------------------------------------
# PASO 2: importar staging -> oficios (INTERNO)
# --------------------------------------------------------------------
def importar_a_oficios(conn, dry_run: bool = False) -> None:
    with conn.cursor() as cur:
        # ids clave
        cur.execute("SELECT id FROM tipos_oficio WHERE clave='INTERNO'")
        tipo_interno_id = cur.fetchone()[0]

        cur.execute("SELECT id FROM estados_oficio WHERE upper(nombre)='ARCHIVADO' LIMIT 1")
        row = cur.fetchone()
        if not row:
            print("ERROR: no existe estado 'Archivado' en estados_oficio.")
            return
        estado_archivado_id = row[0]

        cur.execute("SELECT id FROM estados_oficio WHERE upper(nombre)='EN PROCESO' LIMIT 1")
        row = cur.fetchone()
        estado_en_proceso_id = row[0] if row else estado_archivado_id

        cur.execute("SELECT id FROM usuarios WHERE lower(username)=%s LIMIT 1", (USUARIO_CAPTURO,))
        row = cur.fetchone()
        if not row:
            print(f"WARNING: no existe usuario '{USUARIO_CAPTURO}'. usuario_capturo_id se dejará NULL.")
            usuario_capturo_id = None
        else:
            usuario_capturo_id = row[0]

        # Leer staging pendiente
        cur.execute("""SELECT id, fila_excel, fecha_recibido, folio_minutario,
                              nombre_direccion, folio_direccion, asunto,
                              fecha_oficio_tics, folio_oficio_tics, realizo,
                              acuse_recibido, status, capturo
                         FROM import_folios_interno_sub_raw
                        WHERE procesado = FALSE
                        ORDER BY fila_excel""")
        rows = cur.fetchall()

        area_cache: dict = {}
        ok = 0
        errores = 0

        # Cálculo de numero_folio para INTERNOS:
        # usamos un espacio propio (>=10000) para NO chocar con EXTERNOS cuyo
        # numero_folio suele ser 1..9999. Dejamos INTERNO en 10000+consecutivo.
        cur.execute("""SELECT COALESCE(MAX(numero_folio),9999)
                         FROM oficios
                        WHERE tipo_oficio_id = %s""", (tipo_interno_id,))
        max_interno = cur.fetchone()[0]
        if max_interno < 10000:
            max_interno = 10000  # base para internos
        consecutivo = max_interno

        for r in rows:
            (sid, fila_excel, fecha_recibido, folio_minutario, nombre_direccion,
             folio_direccion, asunto, fecha_oficio_tics, folio_oficio_tics,
             realizo, acuse_recibido, status, capturo) = r

            try:
                if not asunto or not fecha_recibido:
                    raise ValueError(f"Falta asunto o fecha_recibido (fila {fila_excel})")

                area_id = resolver_area(nombre_direccion, area_cache, cur)
                if area_id is None:
                    raise ValueError(f"No se pudo resolver area interna: {nombre_direccion!r}")

                # mapear STATUS
                status_up = (status or "").strip().upper()
                if status_up == "ARCHIVADO":
                    estado_id = estado_archivado_id
                elif status_up == "":
                    estado_id = estado_en_proceso_id
                else:
                    estado_id = estado_en_proceso_id  # por defecto si texto no reconocido

                consecutivo += 1
                anio_folio = int(fecha_recibido[:4]) if fecha_recibido else datetime.now().year

                if dry_run:
                    print(f"  [DRY] fila {fila_excel}  area={area_id}  "
                          f"folio_int='{folio_oficio_tics}'  asunto='{asunto[:40]}'")
                    ok += 1
                    continue

                cur.execute("""
                    INSERT INTO oficios
                        (numero_folio, anio_folio, folio_minutario, folio_direccion,
                         folio_interno_texto,
                         tipo_oficio_id, area_interna_id, dependencia_id,
                         estado_id, asunto, observaciones,
                         fecha_recepcion, fecha_oficio_tics, fecha_acuse,
                         realizo, usuario_capturo_id)
                    VALUES
                        (%s, %s, %s, %s,
                         %s,
                         %s, %s, NULL,
                         %s, %s, NULL,
                         %s, %s, %s,
                         %s, %s)
                    RETURNING id
                """, (
                    consecutivo, anio_folio,
                    folio_minutario, folio_direccion,
                    folio_oficio_tics,
                    tipo_interno_id, area_id,
                    estado_id, asunto,
                    fecha_recibido, fecha_oficio_tics, acuse_recibido,
                    (realizo or capturo), usuario_capturo_id
                ))
                new_id = cur.fetchone()[0]

                cur.execute("""UPDATE import_folios_interno_sub_raw
                                  SET procesado = TRUE, oficio_id = %s, error_mensaje = NULL
                                WHERE id = %s""", (new_id, sid))
                ok += 1
            except Exception as e:
                errores += 1
                conn.rollback()
                with conn.cursor() as cerr:
                    cerr.execute("""UPDATE import_folios_interno_sub_raw
                                       SET error_mensaje = %s
                                     WHERE id = %s""", (str(e), sid))
                    conn.commit()
                print(f"  [ERROR fila {fila_excel}] {e}")

        if not dry_run:
            conn.commit()
        print(f"Resultados: OK={ok}  Errores={errores}  Total staging={len(rows)}")

# --------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--staging", action="store_true", help="Carga SOLO la tabla staging.")
    ap.add_argument("--dry-run", action="store_true", help="Simula importación sin escribir en oficios.")
    ap.add_argument("--importar", action="store_true", help="Importa staging a oficios (INTERNO).")
    args = ap.parse_args()

    if not (args.staging or args.dry_run or args.importar):
        ap.print_help()
        sys.exit(0)

    conn = pg_connect()
    try:
        if args.staging:
            llenar_staging(conn)
        if args.dry_run:
            importar_a_oficios(conn, dry_run=True)
        if args.importar:
            importar_a_oficios(conn, dry_run=False)
    finally:
        conn.close()

if __name__ == "__main__":
    main()
