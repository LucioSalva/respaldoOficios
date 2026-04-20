"""
Inspector de la hoja OFICIOS DE CONOCIMIENTO (BASE GENERAL.xlsx).

Objetivo: mapear columnas reales, tipos de dato, filas utiles, duplicados
y comparar con BASE GENERAL (hoja oficios externos) para definir la
estrategia de importacion a import_oficios_conocimiento_raw.

No escribe nada a BD: solo inspeccion y reporte.
"""
from __future__ import annotations
import os
import unicodedata
from collections import Counter
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "BASE GENERAL.xlsx")
SHEET = "OFICIOS DE CONOCIMIENTO"


def sa(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET}' no existe. Disponibles: {wb.sheetnames}")
        return
    ws = wb[SHEET]
    print(f"Hoja: {ws.title} | max_row={ws.max_row} | max_col={ws.max_column}")

    # Escanear primeras filas para detectar cabecera
    print("\n--- PRIMERAS 8 FILAS (para detectar cabecera) ---")
    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    for i, r in enumerate(rows, start=1):
        cells = [fmt(v) for v in r]
        print(f"Fila {i}: " + " | ".join(cells[:20]))

    # Detectar fila de cabecera (la que tiene mas strings)
    header_row_idx = None
    header = None
    for i, r in enumerate(rows, start=1):
        text_count = sum(1 for v in r if isinstance(v, str) and v.strip())
        if text_count >= 4 and text_count > (len(r) or 0) * 0.3:
            header_row_idx = i
            header = [fmt(v).strip() for v in r]
            break
    if header is None:
        header_row_idx = 1
        header = [fmt(v).strip() for v in rows[0]] if rows else []

    # Recortar columnas vacias al final
    while header and header[-1] == "":
        header.pop()
    print(f"\nCabecera detectada en fila {header_row_idx}:")
    for idx, h in enumerate(header):
        print(f"  [{idx}] {h!r}")

    # Leer todo el dataset
    data_rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None:
            continue
        # skip si todas nulas
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        data_rows.append(row)
    print(f"\nFilas con datos: {len(data_rows)}")

    # Perfilado por columna
    n_cols = len(header)
    print("\n--- PERFIL POR COLUMNA (primeras 20 columnas) ---")
    for c in range(min(n_cols, 20)):
        values = [r[c] if c < len(r) else None for r in data_rows]
        non_null = [v for v in values if v is not None and (not isinstance(v, str) or v.strip())]
        sample = [fmt(v) for v in non_null[:5]]
        types = Counter(type(v).__name__ for v in non_null)
        print(f"  [{c}] {header[c]!r:45s} llenas={len(non_null):4d}/{len(data_rows):4d}  tipos={dict(types)}")
        print(f"        ejemplos: {sample}")

    # Analisis especifico: estatus / status / dependencia / area / fechas
    print("\n--- VALORES UNICOS DE COLUMNAS CANDIDATAS ---")
    # buscar col de estatus
    for target_kw in ["STATUS", "ESTATUS", "ESTADO"]:
        for c, h in enumerate(header):
            if target_kw in sa(h):
                vals = Counter(fmt(r[c]).strip() for r in data_rows if c < len(r))
                print(f"\n  Columna [{c}] '{h}' top valores:")
                for v, n in vals.most_common(15):
                    print(f"    {n:4d}x  {v!r}")
                break
    # buscar col de dependencia / direccion
    for target_kw in ["DEPENDENCIA", "DIRECCION", "AREA", "DE LA DIRECCION", "QUIEN LO REMITE"]:
        for c, h in enumerate(header):
            if target_kw in sa(h):
                vals = Counter(fmt(r[c]).strip() for r in data_rows if c < len(r))
                print(f"\n  Columna [{c}] '{h}' ({target_kw}) top 10:")
                for v, n in vals.most_common(10):
                    print(f"    {n:4d}x  {v!r}")
                break

    # Duplicados por folio/asunto/fecha
    print("\n--- DUPLICADOS POTENCIALES ---")
    # detectar columna folio
    folio_col = None
    asunto_col = None
    fecha_col = None
    for c, h in enumerate(header):
        ha = sa(h)
        if folio_col is None and ("FOLIO" in ha):
            folio_col = c
        if asunto_col is None and "ASUNTO" in ha:
            asunto_col = c
        if fecha_col is None and ("FECHA" in ha and "RECIB" in ha):
            fecha_col = c
    print(f"  folio_col={folio_col} asunto_col={asunto_col} fecha_col={fecha_col}")
    if folio_col is not None:
        folios = Counter(fmt(r[folio_col]).strip() for r in data_rows if folio_col < len(r) and r[folio_col])
        dup = {k: v for k, v in folios.items() if v > 1}
        print(f"  Folios duplicados (exactos, no vacios): {len(dup)}")
        for k, v in list(dup.items())[:10]:
            print(f"    {v}x  {k!r}")

    # Comparar cabeceras con hoja BASE GENERAL (oficios externos)
    print("\n--- COMPARACION CON HOJA 'BASE GENERAL' (oficios externos) ---")
    if "BASE GENERAL " in wb.sheetnames or "BASE GENERAL" in wb.sheetnames:
        ws_bg = wb["BASE GENERAL " if "BASE GENERAL " in wb.sheetnames else "BASE GENERAL"]
        bg_rows = list(ws_bg.iter_rows(min_row=1, max_row=8, values_only=True))
        for i, r in enumerate(bg_rows, start=1):
            cells = [fmt(v) for v in r]
            text_count = sum(1 for v in r if isinstance(v, str) and v.strip())
            if text_count >= 4:
                print(f"  Cabecera BASE GENERAL fila {i}: {[c for c in cells if c][:20]}")
                break

    print("\n--- FIN INSPECCION ---")


if __name__ == "__main__":
    main()
