"""
Importador FASE 3 — VACACIONES SINDICALIZADOS 2026.

Archivo fuente exacto:
    BASE VACACIONES 2026.xlsx
Hoja fuente exacta:
    VACACIONES SINDICALIZADOS 2026

Columnas (14):
    1 CONS. | 2 FECHA DE VACIONES (sic) | 3 NO. DE EMPLEADO | 4 NOMBRE |
    5 S | 6 C | 7 DIAS POR JUSTIFICAR (texto libre) | 8 PERIODO POR JUSTIFICAR |
    9 DIAS QUE CORRESPONDEN | 10 DIAS OTORGADOS | 11 DIAS PENDIENTES |
    12 REGRESA A LABORAR | 13 RECIBIDO POR LA COORDINACION | 14 FOLIO TM

Flujo:
    1. Validar archivo y hoja exactos.
    2. TRUNCATE import_vacaciones_sindicalizados_2026_raw.
    3. Insertar 1 staging row por cada fila del Excel (datos 2..70).
    4. Resolver/crear personal por num_empleado -> nombre_normalizado.
    5. Marcar tipo_personal = SINDICALIZADO si S lleno y C vacio;
       ambiguo -> estado_revision=PENDIENTE_REVISION.
    6. Parsear fechas tolerante (datetime/string/None).
    7. Insertar movimientos (estatus derivado); si S/C ambiguo o fechas
       no parseables -> PENDIENTE_REVISION.
    8. Recalcular saldos por (personal_id, periodo_id):
        dias_asignados = MAX(dias_corresponden en el grupo),
        dias_usados    = SUM(dias_tomados) excluyendo CANCELADA y PENDIENTE_REVISION,
        tiene_inconsistencia = (dias_pendientes_excel ultimo-movimiento
                                <> dias_asignados - dias_usados).
    9. Resumen final en stdout.

Idempotente: la migracion 11 ya dejo BD limpia. Re-ejecutarlo TRUNCATEA staging
y borra SOLO los movimientos cuyo origen='IMPORT_2026' (no toca manuales).
Los saldos se recalculan para cada par (personal_id, periodo_id) afectado.

Uso:
    python tools/importar_vacaciones_sindicalizados_2026.py
    python tools/importar_vacaciones_sindicalizados_2026.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, date
from typing import Any, Optional

try:
    import psycopg2
    from psycopg2.extras import Json, DictCursor
except ImportError:
    print("ERROR: falta psycopg2. pip install psycopg2-binary"); sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. pip install openpyxl"); sys.exit(1)


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(ROOT, "BASE VACACIONES 2026.xlsx")
SHEET_NAME = "VACACIONES SINDICALIZADOS 2026"

DB = {
    "host":     os.getenv("DB_HOST", "host.docker.internal"),
    "port":     int(os.getenv("DB_PORT", "5432")),
    "dbname":   os.getenv("DB_NAME", "respaldooficios"),
    "user":     os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASS", "admin"),
}

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
DIAS_SEMANA = (
    "lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"
)


# ---------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------
def strip_accents(s: str) -> str:
    if s is None:
        return ""
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def norm_nombre(s: str) -> str:
    if not s:
        return ""
    s = strip_accents(s).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def cell_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def to_date(v: Any) -> Optional[date]:
    """Parseo tolerante de fecha. Devuelve None si no se puede."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # ISO
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # "LUNES 5 DE ENERO DE 2026" / "5 DE ENERO DE 2026" / "5 DE ENERO 2026"
    low = strip_accents(s).lower()
    for dia in DIAS_SEMANA:
        if low.startswith(dia):
            low = low[len(dia):].strip()
            break
    m = re.search(r"(\d{1,2})\s+de\s+([a-z]+)(?:\s+de)?\s+(\d{4})", low)
    if m:
        d = int(m.group(1)); mes = MESES_ES.get(m.group(2)); anio = int(m.group(3))
        if mes and 1 <= d <= 31:
            try:
                return date(anio, mes, d)
            except ValueError:
                return None
    return None


def periodo_clave(raw: str) -> Optional[tuple[str, str, int, int]]:
    """
    Devuelve (clave, nombre, anio, orden) a partir del texto del Excel.
    '1ER PERIODO 2025' -> ('1ER_PERIODO_2025', '1er periodo 2025', 2025, 1)
    '2DO PERIODO 2025' -> ('2DO_PERIODO_2025', '2do periodo 2025', 2025, 2)
    Soporta 1ER/1RO/2DO/PRIMER/SEGUNDO.
    """
    if not raw:
        return None
    s = strip_accents(str(raw)).upper()
    s = re.sub(r"\s+", " ", s).strip()
    m_anio = re.search(r"(20\d{2})", s)
    if not m_anio:
        return None
    anio = int(m_anio.group(1))
    if re.search(r"\b(1ER|1RO|PRIMER|PRIMERO)\b", s):
        orden = 1; pref = "1ER"
    elif re.search(r"\b(2DO|SEGUNDO|2DA)\b", s):
        orden = 2; pref = "2DO"
    elif re.search(r"\b(3ER|TERCER|3RO)\b", s):
        orden = 3; pref = "3ER"
    elif re.search(r"\b(4TO|CUARTO)\b", s):
        orden = 4; pref = "4TO"
    else:
        return None
    clave = f"{pref}_PERIODO_{anio}"
    nombre = f"{pref.lower()} periodo {anio}"
    return clave, nombre, anio, orden


def tipo_personal_id(cur, clave: str) -> Optional[int]:
    cur.execute("SELECT id FROM tipos_personal WHERE clave = %s", (clave,))
    r = cur.fetchone()
    return r[0] if r else None


def estatus_id_by(cur, clave: str) -> int:
    cur.execute("SELECT id FROM estatus_vacaciones WHERE clave = %s", (clave,))
    r = cur.fetchone()
    if not r:
        raise RuntimeError(f"estatus_vacaciones.{clave} no existe")
    return r[0]


def get_or_create_periodo(cur, raw: str, stats: dict) -> Optional[int]:
    parsed = periodo_clave(raw)
    if not parsed:
        return None
    clave, nombre, anio, orden = parsed
    cur.execute("SELECT id FROM vacaciones_periodos WHERE clave = %s", (clave,))
    r = cur.fetchone()
    if r:
        return r[0]
    cur.execute(
        "INSERT INTO vacaciones_periodos (clave, nombre, anio, orden) "
        "VALUES (%s, %s, %s, %s) RETURNING id",
        (clave, nombre, anio, orden),
    )
    stats.setdefault("periodos_creados", []).append(clave)
    return cur.fetchone()[0]


def resolver_personal(
    cur, num_empleado: Optional[str], nombre: str, tipo_id: int
) -> tuple[Optional[int], str, str]:
    """
    Devuelve (personal_id, accion, motivo).
    accion in {CREADO, ACTUALIZADO, ENCONTRADO, OMITIDO}.
    """
    nombre = (nombre or "").strip()
    if not nombre:
        return None, "OMITIDO", "Nombre vacio"
    norm = norm_nombre(nombre)
    # 1) por numero_empleado
    if num_empleado:
        cur.execute(
            "SELECT id FROM personal WHERE numero_empleado = %s LIMIT 1",
            (num_empleado,),
        )
        r = cur.fetchone()
        if r:
            return r[0], "ENCONTRADO", "match por numero_empleado"
    # 2) por nombre_normalizado
    cur.execute(
        "SELECT id, numero_empleado FROM personal "
        "WHERE nombre_normalizado = %s LIMIT 1",
        (norm,),
    )
    r = cur.fetchone()
    if r:
        pid, num_db = r[0], r[1]
        if num_db and num_empleado and num_db != num_empleado:
            return pid, "OMITIDO", (
                f"Mismo nombre pero num_empleado distinto "
                f"(BD={num_db}, Excel={num_empleado})"
            )
        if not num_db and num_empleado:
            cur.execute(
                "UPDATE personal SET numero_empleado = %s WHERE id = %s",
                (num_empleado, pid),
            )
            return pid, "ACTUALIZADO", "num_empleado rellenado"
        return pid, "ENCONTRADO", "match por nombre"
    # 3) crear
    cur.execute(
        "INSERT INTO personal (numero_empleado, nombre_completo, "
        "nombre_normalizado, tipo_personal_id) "
        "VALUES (%s, %s, %s, %s) RETURNING id",
        (num_empleado or None, nombre, norm, tipo_id),
    )
    return cur.fetchone()[0], "CREADO", "personal creado"


def derivar_estatus_clave(
    fini: Optional[date], ffin: Optional[date], freg: Optional[date],
    hoy: date, pendiente_revision: bool
) -> str:
    if pendiente_revision:
        return "PENDIENTE_REVISION"
    if freg and freg <= hoy:
        return "FINALIZADA"
    if ffin and ffin < hoy:
        return "FINALIZADA"
    if fini and fini > hoy:
        return "PROGRAMADA"
    if fini and ffin and fini <= hoy <= ffin:
        return "ACTIVA"
    # Si no hay fechas, lo mas conservador es PROGRAMADA
    return "PROGRAMADA"


# ---------------------------------------------------------------------
# Core
# ---------------------------------------------------------------------
def leer_filas_excel() -> list[dict]:
    if not os.path.isfile(XLSX_PATH):
        raise FileNotFoundError(f"Archivo no encontrado: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Hoja '{SHEET_NAME}' no existe en {XLSX_PATH}. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        # Columnas 1..14 (indices 0..13) segun FASE 1
        row = list(row) + [None] * (14 - len(row))
        row = row[:14]
        # Saltar filas completamente vacias
        if all((c is None or str(c).strip() == "") for c in row):
            continue
        d = {
            "fila_excel": idx,
            "cons":                row[0],
            "fecha_vaciones":      row[1],
            "numero_empleado":     row[2],
            "nombre":              row[3],
            "s":                   row[4],
            "c":                   row[5],
            "dias_por_justificar": row[6],
            "periodo_por_justificar": row[7],
            "dias_corresponden":   row[8],
            "dias_otorgados":      row[9],
            "dias_pendientes":     row[10],
            "regresa_a_laborar":   row[11],
            "recibido_coord":      row[12],
            "folio_tm":            row[13],
        }
        # Si no hay ni numero_empleado ni nombre, saltar
        if not cell_text(d["numero_empleado"]) and not cell_text(d["nombre"]):
            continue
        filas.append(d)
    return filas


def importar(dry_run: bool = False) -> int:
    print(f"[INFO] Abriendo {XLSX_PATH}")
    print(f"[INFO] Hoja: {SHEET_NAME}")
    filas = leer_filas_excel()
    print(f"[INFO] Filas detectadas: {len(filas)}")

    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    stats = {
        "filas_excel": len(filas),
        "staging_insertados": 0,
        "personal_creado": 0,
        "personal_actualizado": 0,
        "personal_encontrado": 0,
        "personal_omitido": 0,
        "movimientos_creados": 0,
        "pendientes_revision": 0,
        "periodos_creados": [],
        "saldos_afectados": 0,
        "saldos_con_inconsistencia": 0,
        "errores": [],
    }

    try:
        with conn.cursor() as cur:
            # Limpieza idempotente
            cur.execute("TRUNCATE TABLE import_vacaciones_sindicalizados_2026_raw RESTART IDENTITY")
            cur.execute(
                "DELETE FROM vacaciones_movimientos WHERE origen = 'IMPORT_2026'"
            )
            tipo_sind = tipo_personal_id(cur, "SINDICALIZADO")
            tipo_conf = tipo_personal_id(cur, "CONFIANZA")
            if not tipo_sind:
                raise RuntimeError("tipos_personal.SINDICALIZADO no existe")
            est_programada = estatus_id_by(cur, "PROGRAMADA")
            est_activa     = estatus_id_by(cur, "ACTIVA")
            est_finalizada = estatus_id_by(cur, "FINALIZADA")
            est_pendrev    = estatus_id_by(cur, "PENDIENTE_REVISION")
            mapa_estatus = {
                "PROGRAMADA":         est_programada,
                "ACTIVA":             est_activa,
                "FINALIZADA":         est_finalizada,
                "PENDIENTE_REVISION": est_pendrev,
            }
            hoy = date.today()
            pares_afectados: set[tuple[int, int]] = set()

            for f in filas:
                num_emp   = cell_text(f["numero_empleado"]).split(".")[0] or None
                nombre    = cell_text(f["nombre"]).strip() or None
                s_val     = cell_text(f["s"]).strip()
                c_val     = cell_text(f["c"]).strip()
                periodo_raw = cell_text(f["periodo_por_justificar"]).strip()
                folio_tm  = cell_text(f["folio_tm"]).strip() or None
                dias_corr = to_int(f["dias_corresponden"]) or 0
                dias_otor = to_int(f["dias_otorgados"])    or 0
                dias_pend = to_int(f["dias_pendientes"])
                fecha_vac = to_date(f["fecha_vaciones"])
                freg_raw  = cell_text(f["regresa_a_laborar"])
                freg      = to_date(f["regresa_a_laborar"])
                frec_coord = to_date(f["recibido_coord"])

                observ_parts = []
                estado_revision = "OK"
                error_msg = None

                # Tipo por S/C (D6)
                if s_val and not c_val:
                    tipo_id = tipo_sind
                    tipo_resuelto = "SINDICALIZADO"
                elif c_val and not s_val:
                    tipo_id = tipo_conf or tipo_sind  # fallback, aunque no deberia ocurrir aqui
                    tipo_resuelto = "CONFIANZA"
                    estado_revision = "PENDIENTE_REVISION"
                    error_msg = "marca C presente en hoja solo-sindicalizados"
                    observ_parts.append(error_msg)
                else:
                    tipo_id = tipo_sind
                    tipo_resuelto = "SINDICALIZADO"
                    if s_val and c_val:
                        estado_revision = "PENDIENTE_REVISION"
                        error_msg = "PENDIENTE_REVISION: marca S/C ambigua"
                        observ_parts.append(error_msg)
                    elif not s_val and not c_val:
                        estado_revision = "PENDIENTE_REVISION"
                        error_msg = "PENDIENTE_REVISION: sin marca S ni C"
                        observ_parts.append(error_msg)

                # Fechas no parseables -> observar pero no romper
                if cell_text(f["fecha_vaciones"]) and fecha_vac is None:
                    observ_parts.append(f"fecha_vacaciones no parseable: {cell_text(f['fecha_vaciones'])}")
                if freg_raw and freg is None:
                    observ_parts.append(f"regreso no parseable: {freg_raw}")

                # Periodo
                periodo_id = get_or_create_periodo(cur, periodo_raw, stats)
                if not periodo_id:
                    estado_revision = "PENDIENTE_REVISION"
                    error_msg = (error_msg or "") + " | periodo no reconocido"
                    observ_parts.append(f"periodo no reconocido: {periodo_raw}")

                # Personal
                if nombre:
                    pid, accion_p, motivo_p = resolver_personal(
                        cur, num_emp, nombre, tipo_id
                    )
                    if accion_p == "CREADO":
                        stats["personal_creado"] += 1
                    elif accion_p == "ACTUALIZADO":
                        stats["personal_actualizado"] += 1
                    elif accion_p == "ENCONTRADO":
                        stats["personal_encontrado"] += 1
                    else:
                        stats["personal_omitido"] += 1
                        estado_revision = "PENDIENTE_REVISION"
                        error_msg = (error_msg or "") + f" | personal: {motivo_p}"
                        observ_parts.append(f"personal: {motivo_p}")
                else:
                    pid = None
                    stats["personal_omitido"] += 1
                    estado_revision = "PENDIENTE_REVISION"
                    error_msg = (error_msg or "") + " | nombre vacio"

                # Staging insert
                raw_data = {k: (v.isoformat() if isinstance(v, (datetime, date)) else v)
                            for k, v in f.items()}
                cur.execute(
                    """
                    INSERT INTO import_vacaciones_sindicalizados_2026_raw
                        (source_file, source_sheet, fila_excel, raw_data,
                         cons_raw, fecha_vaciones_raw, numero_empleado_raw,
                         nombre_raw, s_raw, c_raw, dias_por_justificar_raw,
                         periodo_por_justificar_raw, dias_corresponden_raw,
                         dias_otorgados_raw, dias_pendientes_raw,
                         regresa_a_laborar_raw, recibido_por_coordinacion_raw,
                         folio_tm_raw,
                         personal_id, periodo_id, tipo_personal_resuelto,
                         estado_revision, error_importacion)
                    VALUES (%s,%s,%s,%s,
                            %s,%s,%s,
                            %s,%s,%s,%s,
                            %s,%s,
                            %s,%s,
                            %s,%s,
                            %s,
                            %s,%s,%s,
                            %s,%s)
                    RETURNING id
                    """,
                    (
                        os.path.basename(XLSX_PATH), SHEET_NAME, f["fila_excel"],
                        Json(raw_data),
                        cell_text(f["cons"]),
                        cell_text(f["fecha_vaciones"]),
                        cell_text(f["numero_empleado"]),
                        cell_text(f["nombre"]),
                        s_val, c_val,
                        cell_text(f["dias_por_justificar"]),
                        periodo_raw,
                        cell_text(f["dias_corresponden"]),
                        cell_text(f["dias_otorgados"]),
                        cell_text(f["dias_pendientes"]),
                        freg_raw,
                        cell_text(f["recibido_coord"]),
                        cell_text(f["folio_tm"]),
                        pid, periodo_id, tipo_resuelto,
                        estado_revision, error_msg,
                    ),
                )
                staging_id = cur.fetchone()[0]
                stats["staging_insertados"] += 1

                # Si no hay personal o periodo -> no se crea movimiento
                if not pid or not periodo_id:
                    stats["pendientes_revision"] += 1
                    continue

                # Estatus
                estatus_clave = derivar_estatus_clave(
                    None, None, freg, hoy,
                    pendiente_revision=(estado_revision == "PENDIENTE_REVISION"),
                )
                if estatus_clave == "PENDIENTE_REVISION":
                    stats["pendientes_revision"] += 1

                # Insertar movimiento
                cur.execute(
                    """
                    INSERT INTO vacaciones_movimientos
                        (personal_id, periodo_id,
                         fecha_vacaciones, fecha_inicio, fecha_fin, fecha_regreso,
                         fecha_recibido_coord,
                         dias_corresponden, dias_tomados, dias_pendientes_excel,
                         folio_tm_vacaciones,
                         estatus_id, capturado_por_user_id, observaciones,
                         origen, import_raw_id)
                    VALUES (%s,%s,
                            %s,%s,%s,%s,
                            %s,
                            %s,%s,%s,
                            %s,
                            %s,NULL,%s,
                            'IMPORT_2026', %s)
                    RETURNING id
                    """,
                    (
                        pid, periodo_id,
                        fecha_vac, None, None, freg,
                        frec_coord,
                        dias_corr, dias_otor, dias_pend,
                        folio_tm,
                        mapa_estatus[estatus_clave],
                        " | ".join(observ_parts) if observ_parts else None,
                        staging_id,
                    ),
                )
                mov_id = cur.fetchone()[0]
                cur.execute(
                    "UPDATE import_vacaciones_sindicalizados_2026_raw "
                    "SET movimiento_id = %s, importado = TRUE WHERE id = %s",
                    (mov_id, staging_id),
                )
                stats["movimientos_creados"] += 1
                pares_afectados.add((pid, periodo_id))

            # Recalcular saldos (D1)
            for pid, per_id in pares_afectados:
                # pg_advisory_xact_lock por par para evitar race condition
                cur.execute(
                    "SELECT pg_advisory_xact_lock(%s, %s)", (pid, per_id)
                )
                cur.execute(
                    """
                    SELECT
                        COALESCE(MAX(dias_corresponden), 0)                              AS dias_asignados,
                        COALESCE(SUM(CASE WHEN ev.clave IN ('CANCELADA','PENDIENTE_REVISION')
                                          THEN 0 ELSE m.dias_tomados END), 0)             AS dias_usados,
                        (SELECT m2.dias_pendientes_excel
                           FROM vacaciones_movimientos m2
                           JOIN estatus_vacaciones ev2 ON ev2.id = m2.estatus_id
                          WHERE m2.personal_id = %s AND m2.periodo_id = %s
                            AND ev2.clave NOT IN ('CANCELADA','PENDIENTE_REVISION')
                          ORDER BY m2.id DESC LIMIT 1)                                     AS dpe_ultimo
                      FROM vacaciones_movimientos m
                      JOIN estatus_vacaciones ev ON ev.id = m.estatus_id
                     WHERE m.personal_id = %s AND m.periodo_id = %s
                    """,
                    (pid, per_id, pid, per_id),
                )
                row = cur.fetchone()
                d_asig, d_usa, dpe_ult = (row[0] or 0, row[1] or 0, row[2])
                derivado_restante = d_asig - d_usa
                inc = False
                obs = None
                if dpe_ult is not None and dpe_ult != derivado_restante:
                    inc = True
                    obs = (
                        f"Inconsistencia: DIAS_PENDIENTES en Excel={dpe_ult} "
                        f"difiere de (asignados {d_asig} - usados {d_usa}) = {derivado_restante}."
                    )

                cur.execute(
                    """
                    INSERT INTO vacaciones_saldos
                        (personal_id, periodo_id, dias_asignados, dias_usados,
                         tiene_inconsistencia, fuente, observaciones)
                    VALUES (%s, %s, %s, %s, %s, 'DERIVADO_IMPORT_2026', %s)
                    ON CONFLICT (personal_id, periodo_id)
                    DO UPDATE SET
                        dias_asignados       = EXCLUDED.dias_asignados,
                        dias_usados          = EXCLUDED.dias_usados,
                        tiene_inconsistencia = EXCLUDED.tiene_inconsistencia,
                        fuente               = EXCLUDED.fuente,
                        observaciones        = EXCLUDED.observaciones,
                        updated_at           = NOW()
                    """,
                    (pid, per_id, d_asig, d_usa, inc, obs),
                )
                stats["saldos_afectados"] += 1
                if inc:
                    stats["saldos_con_inconsistencia"] += 1

        if dry_run:
            conn.rollback()
            print("[INFO] --dry-run -> ROLLBACK aplicado, nada persistido.")
        else:
            conn.commit()
            print("[INFO] COMMIT OK.")
    except Exception as e:
        conn.rollback()
        print(f"[ERROR] Rollback por: {e}", file=sys.stderr)
        stats["errores"].append(str(e))
        conn.close()
        print_resumen(stats)
        return 2
    conn.close()
    print_resumen(stats)
    return 0


def print_resumen(stats: dict) -> None:
    print("\n===== RESUMEN IMPORTACION VACACIONES SINDICALIZADOS 2026 =====")
    for k in (
        "filas_excel", "staging_insertados",
        "personal_creado", "personal_actualizado", "personal_encontrado",
        "personal_omitido",
        "movimientos_creados", "pendientes_revision",
        "saldos_afectados", "saldos_con_inconsistencia",
    ):
        print(f"  {k:30s}: {stats.get(k)}")
    if stats.get("periodos_creados"):
        print(f"  periodos creados dinamicos  : {stats['periodos_creados']}")
    if stats.get("errores"):
        print("  errores:")
        for e in stats["errores"]:
            print(f"    - {e}")
    print("==============================================================\n")


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--dry-run", action="store_true", help="No hacer COMMIT.")
    args = p.parse_args()
    return importar(dry_run=args.dry_run)


if __name__ == "__main__":
    sys.exit(main())
