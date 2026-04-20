"""
Inspeccion de hoja de INCIDENCIAS en los Excels de la carpeta raiz.
Busca automaticamente hojas cuyo nombre contenga 'INCID' (INCIDENCIAS, CONTROL INCIDENCIAS, etc).
Reporta:
  - Hojas candidatas encontradas (archivo + nombre exacto).
  - Encabezados de la fila 1.
  - Conteo de filas con datos.
  - Tipos distintos detectados (columna TIPO).
  - Estatus distintos detectados.
  - Fechas invalidas.
  - Empleados sin num_empleado.
  - Duplicados aproximados.
  - Muestras (primeras 8 filas).

Uso:
    python tools/_inspect_incidencias.py
"""
from __future__ import annotations
import os
import sys
import unicodedata
from collections import Counter
from datetime import datetime, date
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CANDIDATOS = [
    "BASE GENERAL.xlsx",
    "BASE LUCIO.xlsx",
    "BASE VACACIONES 2026.xlsx",
]


def sa(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()


def val(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v
    s = str(v).strip()
    return s if s else None


def inspect_sheet(path, sheet_name):
    full = os.path.join(ROOT, path)
    wb = load_workbook(full, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print("\n" + "=" * 80)
    print(f"ARCHIVO: {path}")
    print(f"HOJA   : {sheet_name!r}")
    print(f"Filas totales (inc. encabezado): {len(rows)}")

    if not rows:
        wb.close()
        return

    header = rows[0]
    print(f"Columnas ({len(header)}):")
    for i, h in enumerate(header):
        print(f"  [{i:>2}] {h!r}")

    data = [r for r in rows[1:] if not all(v is None for v in r)]
    print(f"Filas con datos: {len(data)}")

    # Muestra
    print("\nMuestra primeras 8 filas con datos:")
    for i, r in enumerate(data[:8], start=1):
        print(f"  {i:>2}: {r}")

    # Intentar detectar columnas TIPO y ESTATUS por cabecera
    headers_norm = [sa(h) for h in header]

    def find_col(patterns):
        for i, h in enumerate(headers_norm):
            for p in patterns:
                if p in h:
                    return i
        return None

    idx_tipo = find_col(["TIPO", "INCIDENCIA"])
    idx_estatus = find_col(["ESTATUS", "ESTADO"])
    idx_fecha = find_col(["FECHA"])
    idx_nombre = find_col(["NOMBRE"])
    idx_num = find_col(["NO EMPLEADO", "NUM EMPLEADO", "NO. EMPLEADO", "N EMPLEADO", "NUMERO EMPLEADO", "NO_EMP", "EMPLEADO"])

    print("\nColumnas detectadas heuristicamente:")
    print(f"  TIPO       -> idx {idx_tipo}")
    print(f"  ESTATUS    -> idx {idx_estatus}")
    print(f"  FECHA      -> idx {idx_fecha}")
    print(f"  NOMBRE     -> idx {idx_nombre}")
    print(f"  NUM_EMP    -> idx {idx_num}")

    # Valores unicos en TIPO
    if idx_tipo is not None:
        tipos = Counter()
        tipos_raw = Counter()
        for r in data:
            v = r[idx_tipo] if idx_tipo < len(r) else None
            if v is None:
                tipos["<NULL>"] += 1
                continue
            s = str(v).strip()
            tipos_raw[s] += 1
            tipos[sa(s)] += 1
        print(f"\nTIPOS distintos (normalizados) = {len(tipos)}")
        for k, c in tipos.most_common():
            print(f"  {c:>4}  {k!r}")
        print(f"\nTIPOS distintos (RAW original) = {len(tipos_raw)}")
        for k, c in tipos_raw.most_common():
            print(f"  {c:>4}  {k!r}")

    # Valores unicos en ESTATUS
    if idx_estatus is not None:
        est = Counter()
        for r in data:
            v = r[idx_estatus] if idx_estatus < len(r) else None
            if v is None:
                est["<NULL>"] += 1
                continue
            est[sa(str(v))] += 1
        print(f"\nESTATUS distintos (normalizados) = {len(est)}")
        for k, c in est.most_common():
            print(f"  {c:>4}  {k!r}")

    # Fechas invalidas / nulas
    sin_fecha = 0
    fechas_raras = 0
    if idx_fecha is not None:
        for r in data:
            v = r[idx_fecha] if idx_fecha < len(r) else None
            if v is None:
                sin_fecha += 1
                continue
            if isinstance(v, (datetime, date)):
                continue
            s = str(v).strip()
            ok = False
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    datetime.strptime(s, fmt); ok = True; break
                except Exception:
                    pass
            if not ok:
                fechas_raras += 1
        print(f"\nFecha nula: {sin_fecha}")
        print(f"Fecha con formato raro: {fechas_raras}")

    # Empleados sin num
    if idx_num is not None:
        sin_num = 0
        for r in data:
            v = r[idx_num] if idx_num < len(r) else None
            if v is None or str(v).strip() == "":
                sin_num += 1
        print(f"Filas sin num_empleado: {sin_num}")

    # Posibles duplicados (fila completa hasheada)
    firmas = Counter()
    for r in data:
        firma = tuple(
            (x.strftime("%Y-%m-%d") if isinstance(x, (datetime, date)) else (str(x).strip() if x is not None else None))
            for x in r
        )
        firmas[firma] += 1
    dup = sum(1 for c in firmas.values() if c > 1)
    print(f"Duplicados exactos (fila completa): {dup}")

    wb.close()


def main():
    candidatos_hojas = []
    for fn in CANDIDATOS:
        full = os.path.join(ROOT, fn)
        if not os.path.exists(full):
            print(f"[{fn}] NO EXISTE, se omite")
            continue
        try:
            wb = load_workbook(full, data_only=True, read_only=True)
        except Exception as e:
            print(f"[{fn}] ERROR abriendo: {e}")
            continue
        print(f"\n[{fn}] hojas: {wb.sheetnames}")
        for sh in wb.sheetnames:
            if "INCID" in sa(sh):
                candidatos_hojas.append((fn, sh))
        wb.close()

    if not candidatos_hojas:
        print("\nNINGUNA hoja con 'INCID' detectada en los Excels. Abortando.")
        sys.exit(0)

    print("\nHOJAS CANDIDATAS DETECTADAS:")
    for fn, sh in candidatos_hojas:
        print(f"  {fn} :: {sh!r}")

    for fn, sh in candidatos_hojas:
        try:
            inspect_sheet(fn, sh)
        except Exception as e:
            print(f"  [{fn} :: {sh}] ERROR inspeccionando: {e}")


if __name__ == "__main__":
    main()
