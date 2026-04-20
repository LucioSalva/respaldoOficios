"""
Analisis de la hoja FOLIOS INTERNO SUB del Excel BASE GENERAL.xlsx.
Solo lectura. No modifica nada.
"""
import sys
import os
from collections import Counter, defaultdict
from openpyxl import load_workbook

XLSX = os.path.join(os.path.dirname(__file__), "..", "BASE GENERAL.xlsx")
XLSX = os.path.abspath(XLSX)
SHEET = "FOLIOS INTERNO SUB"

def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

def is_empty_row(row):
    return all(norm(c) is None for c in row)

def clasif_folio_tics(val):
    if val is None:
        return "VACIO"
    s = str(val).strip().upper()
    if s.startswith("TM/ECA/STIYC"):
        return "FORMATO_TM_ECA_STIYC"
    if s.startswith("SDTICS/"):
        return "FORMATO_SDTICS"
    if "/" in s:
        return "OTRO_CON_SLASH"
    return "LIBRE"

def main():
    print(f"Leyendo: {XLSX}")
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET}' no existe. Disponibles: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET]

    rows = list(ws.iter_rows(values_only=True))
    print(f"Filas totales (incluye encabezado y vacias): {len(rows)}")

    if not rows:
        print("Hoja vacia.")
        return

    header = [norm(c) for c in rows[0]]
    print(f"Encabezados ({len(header)} cols): {header}")

    data = [r for r in rows[1:] if not is_empty_row(r)]
    print(f"Filas con datos (no vacias): {len(data)}")

    idx = {name: i for i, name in enumerate(header) if name}

    def col(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return None
        return norm(row[i])

    # Listado unico de NOMBRE DIRECCION
    dirs = Counter()
    for r in data:
        v = col(r, "NOMBRE DIRECCION") or col(r, "NOMBRE DIRECCIN")
        if v is None:
            # fallback por posicion (columna 2)
            v = norm(r[2]) if len(r) > 2 else None
        if v:
            dirs[str(v).strip().upper()] += 1

    print("\n=== NOMBRE DIRECCION (area interna) distinct + conteo ===")
    for k, v in dirs.most_common():
        print(f"  [{v:3d}]  {k}")
    print(f"  TOTAL distinct areas internas: {len(dirs)}")

    # Distribucion STATUS
    stats = Counter()
    for r in data:
        v = norm(r[9]) if len(r) > 9 else None
        if v:
            stats[str(v).strip().upper()] += 1
        else:
            stats["(VACIO)"] += 1
    print("\n=== STATUS distinct + conteo ===")
    for k, v in stats.most_common():
        print(f"  [{v:3d}]  {k}")

    # Formatos de FOLIO DE OFICIO TICS
    folio_formats = Counter()
    ejemplos = defaultdict(list)
    for r in data:
        v = norm(r[6]) if len(r) > 6 else None
        f = clasif_folio_tics(v)
        folio_formats[f] += 1
        if len(ejemplos[f]) < 5 and v is not None:
            ejemplos[f].append(str(v))
    print("\n=== FORMATO de FOLIO DE OFICIO TICS ===")
    for k, v in folio_formats.most_common():
        print(f"  [{v:3d}]  {k}  ejemplos: {ejemplos[k]}")

    # Columnas vacias totales por columna
    print("\n=== VACIOS POR COLUMNA ===")
    for i, h in enumerate(header):
        vacios = sum(1 for r in data if i >= len(r) or norm(r[i]) is None)
        print(f"  col[{i}] {h!r:40s}  vacios: {vacios}/{len(data)}")

    # Duplicados por FOLIO DE LA DIRECCION
    folio_dir = Counter()
    for r in data:
        v = norm(r[3]) if len(r) > 3 else None
        if v:
            folio_dir[str(v).strip().upper()] += 1
    dups = {k: v for k, v in folio_dir.items() if v > 1}
    print(f"\n=== DUPLICADOS por FOLIO DE LA DIRECCION: {len(dups)} ===")
    for k, v in list(dups.items())[:15]:
        print(f"  [{v}]  {k}")

    # Duplicados por FOLIO DE OFICIO TICS (no vacios)
    folio_tics = Counter()
    for r in data:
        v = norm(r[6]) if len(r) > 6 else None
        if v:
            folio_tics[str(v).strip().upper()] += 1
    dups_t = {k: v for k, v in folio_tics.items() if v > 1}
    print(f"\n=== DUPLICADOS por FOLIO DE OFICIO TICS: {len(dups_t)} ===")
    for k, v in list(dups_t.items())[:15]:
        print(f"  [{v}]  {k}")

    # REALIZO y CAPTURO distinct
    realizo = Counter()
    capturo = Counter()
    for r in data:
        v = norm(r[7]) if len(r) > 7 else None
        if v: realizo[str(v).strip().upper()] += 1
        v = norm(r[10]) if len(r) > 10 else None
        if v: capturo[str(v).strip().upper()] += 1
    print("\n=== REALIZO distinct ===")
    for k, v in realizo.most_common(20):
        print(f"  [{v:3d}]  {k}")
    print("\n=== CAPTURO distinct ===")
    for k, v in capturo.most_common(20):
        print(f"  [{v:3d}]  {k}")

    # Tipos de datos en FOLIO MINUTARIO (col 1) y fechas (col 0, 5, 8)
    print("\n=== MUESTRA DE TIPOS DE DATOS ===")
    for col_idx, col_name in [(0, "FECHA DE RECIBIDO"), (1, "FOLIO MINUTARIO"),
                              (5, "FECHA DE OFICIO TICS"), (8, "ACUSE RECIBIDO")]:
        tipos = Counter()
        muestras = []
        for r in data[:50]:
            v = r[col_idx] if len(r) > col_idx else None
            tipos[type(v).__name__] += 1
            if len(muestras) < 3 and v is not None:
                muestras.append(repr(v))
        print(f"  col[{col_idx}] {col_name}: {dict(tipos)} | muestras: {muestras}")

    # Primeras 3 filas completas de ejemplo
    print("\n=== PRIMERAS 3 FILAS DE DATOS ===")
    for i, r in enumerate(data[:3]):
        print(f"  fila {i+1}: {r}")

    print("\n=== FIN ANALISIS ===")

if __name__ == "__main__":
    main()
