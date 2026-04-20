"""
Importador de INCIDENCIAS.

Fuente real detectada:
  BASE GENERAL.xlsx :: hoja 'INCIDENCIAS'
  Columnas:
    [0]  CONS.
    [1]  FECHA DE INCIDENCIA
    [2]  NO. DE EMPLEADO
    [3]  NOMBRE
    [4]  S/C                          (S = SIND, C = CONF)
    [5]  DIA A JUSTIFICAR              (fecha)
    [6]  E   (marca X = omision entrada)
    [7]  S   (marca X = omision salida)
    [8]  DC  (marca X = dia completo / FALTA)
    [9]  INC (marca X = incidencia generica)
    [10] MOTIVO
    [11] RECIBIDO POR LA COORDINACION  (fecha)

Uso:
    python tools/importar_incidencias.py --staging
    python tools/importar_incidencias.py --dry-run
    python tools/importar_incidencias.py --importar

Reglas:
  - El tipo de incidencia se determina por las columnas E/S/DC/INC (NUNCA
    texto libre). Si ninguna marca esta puesta pero hay motivo => COMISION.
  - NO se crea personal nuevo. Si no hay match por numero_empleado ni por
    nombre_normalizado -> staging.estado_revision='PENDIENTE_REVISION'
    con error_importacion='Personal no encontrado'.
  - UNIQUE (personal_id, tipo, fecha_incid, fecha_ini, fecha_fin) impide
    duplicados. Si choca -> accion='OMITIDO_DUP'.
  - Fechas invalidas en string -> fila va a PENDIENTE_REVISION.
  - Idempotente: --staging hace TRUNCATE de import_incidencias_raw.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from collections import Counter
from datetime import datetime, date
from typing import Any, Optional

try:
    import psycopg2
    from psycopg2.extras import execute_batch, Json
except ImportError:
    print("ERROR: falta psycopg2. pip install psycopg2-binary"); sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. pip install openpyxl"); sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_GEN = os.path.join(ROOT, "BASE GENERAL.xlsx")
SHEET = "INCIDENCIAS"


# ---------------- helpers -----------------

def pg_connect():
    return psycopg2.connect(
        host     = os.environ.get("PGHOST",     "localhost"),
        port     = int(os.environ.get("PGPORT", "5432")),
        dbname   = os.environ.get("PGDATABASE", "respaldooficios"),
        user     = os.environ.get("PGUSER",     "postgres"),
        password = os.environ.get("PGPASSWORD", "admin"),
    )


def sa(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()


def norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    if s.upper() in ("#N/A", "#N/D", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#DIV/0!"):
        return None
    return s


def parse_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"^(\d{1,2}/\d{1,2}/\d{4})\d+$", r"\1", s)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_num_empleado(v: Any) -> Optional[str]:
    """Normaliza num_empleado: upper, sin espacios, quita ceros a la izquierda."""
    s = norm(v)
    if not s:
        return None
    s = s.upper().strip().replace(" ", "")
    s = s.lstrip("0") or "0"
    return s


def mark_x(v: Any) -> bool:
    """Devuelve True si la celda marca 'X' (tolerante a mayusculas y a 'XM' visto en datos)."""
    if v is None:
        return False
    s = str(v).strip().upper()
    return s in ("X", "XX", "XM", "SI", "SÍ", "YES", "1", "TRUE", "V")


def to_jsonable(row: tuple) -> list:
    return [v.strftime("%Y-%m-%d") if isinstance(v, (datetime, date)) else v for v in row]


# -----------------------------------------------------------
# Deteccion de tipo por marcas E/S/DC/INC
# -----------------------------------------------------------
def detectar_tipo(e_mark: bool, s_mark: bool, dc_mark: bool, inc_mark: bool, motivo: Optional[str]) -> Optional[str]:
    """
    Prioridad:
      DC  -> FALTA
      E   -> OMISION_ENTRADA
      S   -> OMISION_SALIDA
      INC -> INCIDENCIA_GENERAL
      Sin marca + motivo con palabras clave -> COMISION / RETARDO / PERMISO / INCAPACIDAD
      Sin marca ni motivo util -> None (staging pendiente)
    """
    if dc_mark:  return "FALTA"
    if e_mark:   return "OMISION_ENTRADA"
    if s_mark:   return "OMISION_SALIDA"
    if inc_mark: return "INCIDENCIA_GENERAL"

    if motivo:
        m = sa(motivo)
        if re.search(r"\bRETARDO", m):                 return "RETARDO"
        if re.search(r"\bINCAPACID", m):               return "INCAPACIDAD"
        if re.search(r"\bPERMISO\b", m):               return "PERMISO"
        if re.search(r"\bDIA ECONOMIC", m):            return "DIA_ECONOMICO"
        if re.search(r"\bJUSTIFICACI", m):             return "JUSTIFICACION"
        # Casos frecuentes en el Excel real: "APOYO EN ...", "REVISION DE ...",
        # "VISITA A ..." => COMISION / apoyo externo
        if re.search(r"\b(APOYO|REVISION|COMISION|VISITA|TRASLADO)\b", m):
            return "COMISION"
    return None


# -----------------------------------------------------------
# FASE 1: STAGING
# -----------------------------------------------------------
def ensure_open(path: str):
    if not os.path.exists(path):
        return None
    return load_workbook(path, data_only=True, read_only=True)


def cargar_staging(conn) -> int:
    wb = ensure_open(XLSX_GEN)
    if not wb:
        print(f"ERROR: no existe {XLSX_GEN}"); return 0
    if SHEET not in wb.sheetnames:
        print(f"ERROR: la hoja {SHEET!r} no existe en {XLSX_GEN}"); return 0

    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))

    rows_all = []
    for idx, r in enumerate(rows[1:], start=2):
        if not r or all(v is None for v in r):
            continue
        # Saltar filas totalmente de plantilla (#N/A)
        num = parse_num_empleado(r[2] if len(r) > 2 else None)
        nombre = norm(r[3] if len(r) > 3 else None)
        if not num and not nombre:
            # Fila vacia / placeholder.
            continue

        rows_all.append((
            "BASE GENERAL.xlsx!INCIDENCIAS",
            idx,
            norm(r[0] if len(r) > 0 else None),      # cons
            parse_date(r[1] if len(r) > 1 else None) or norm(r[1] if len(r) > 1 else None),
            num,
            nombre,
            norm(r[4] if len(r) > 4 else None),      # S/C
            parse_date(r[5] if len(r) > 5 else None) or norm(r[5] if len(r) > 5 else None),
            norm(r[6] if len(r) > 6 else None),      # E
            norm(r[7] if len(r) > 7 else None),      # S
            norm(r[8] if len(r) > 8 else None),      # DC
            norm(r[9] if len(r) > 9 else None),      # INC
            norm(r[10] if len(r) > 10 else None),    # motivo
            parse_date(r[11] if len(r) > 11 else None) or norm(r[11] if len(r) > 11 else None),
            Json(to_jsonable(r)),
        ))

    wb.close()
    if not rows_all:
        print("No hay filas para staging incidencias."); return 0

    with conn.cursor() as cur:
        cur.execute("TRUNCATE import_incidencias_raw RESTART IDENTITY")
        execute_batch(
            cur,
            """INSERT INTO import_incidencias_raw
                (fuente, fila_excel,
                 raw_cons, raw_fecha_incid, raw_num_empleado, raw_nombre,
                 raw_sc, raw_dia_justificar, raw_e, raw_s, raw_dc, raw_inc,
                 raw_motivo, raw_recibido_coord, raw_data)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            rows_all, page_size=100,
        )
        conn.commit()
    print(f"Staging import_incidencias_raw: {len(rows_all)} filas cargadas.")
    return len(rows_all)


# -----------------------------------------------------------
# FASE 2: staging -> incidencias
# -----------------------------------------------------------
def procesar(conn, dry_run: bool = False) -> None:
    stats = Counter()

    with conn.cursor() as cur:
        cur.execute("SELECT id, clave FROM tipos_incidencia")
        mapa_tipos = {clave: tid for (tid, clave) in cur.fetchall()}
        cur.execute("SELECT id, clave FROM estatus_incidencia")
        mapa_est = {clave: eid for (eid, clave) in cur.fetchall()}

        cur.execute(
            """SELECT id, fuente, fila_excel,
                      raw_fecha_incid, raw_num_empleado, raw_nombre,
                      raw_sc, raw_dia_justificar,
                      raw_e, raw_s, raw_dc, raw_inc,
                      raw_motivo, raw_recibido_coord
                 FROM import_incidencias_raw
                WHERE procesado = FALSE
                ORDER BY fila_excel"""
        )
        filas = cur.fetchall()

    est_registrada = mapa_est.get("REGISTRADA")
    est_pend       = mapa_est.get("PENDIENTE_REVISION")

    if not est_registrada or not est_pend:
        print("ERROR: catalogo estatus_incidencia incompleto. Ejecute 07_incidencias.sql.")
        return

    for (sid, fuente, fila_ex,
         raw_fi, raw_num, raw_nombre,
         raw_sc, raw_dj,
         raw_e, raw_s, raw_dc, raw_inc,
         raw_motivo, raw_recib) in filas:

        try:
            stats["total"] += 1
            # Parsear fechas derivadas
            f_incid = parse_date(raw_fi)
            f_just  = parse_date(raw_dj)
            f_recib = parse_date(raw_recib)

            # Detectar tipo
            tipo_clave = detectar_tipo(
                mark_x(raw_e), mark_x(raw_s), mark_x(raw_dc), mark_x(raw_inc),
                raw_motivo,
            )

            # Si no se pudo detectar tipo => PENDIENTE_REVISION
            if not tipo_clave:
                stats["sin_tipo"] += 1
                _marcar_revision(conn, sid, None,
                                 f_incid, f_just, f_just, f_recib,
                                 "No se pudo determinar tipo de incidencia (sin marca E/S/DC/INC y motivo no reconocido).",
                                 None)
                continue

            tipo_id = mapa_tipos.get(tipo_clave)
            if not tipo_id:
                stats["tipo_desconocido"] += 1
                _marcar_revision(conn, sid, None,
                                 f_incid, f_just, f_just, f_recib,
                                 f"Tipo '{tipo_clave}' no existe en catalogo.", tipo_clave)
                continue

            # Resolver personal
            pid = _resolver_personal(conn, raw_num, raw_nombre)
            if not pid:
                stats["personal_no_encontrado"] += 1
                _marcar_revision(conn, sid, None,
                                 f_incid, f_just, f_just, f_recib,
                                 f"Personal no encontrado (num={raw_num!r}, nombre={raw_nombre!r}).",
                                 tipo_clave)
                continue

            # Validar fechas
            if not f_incid and not f_just:
                stats["sin_fecha"] += 1
                _marcar_revision(conn, sid, pid,
                                 None, None, None, f_recib,
                                 "Sin fecha de incidencia ni dia a justificar.", tipo_clave)
                continue

            # Cantidades por tipo
            dias = 1 if tipo_clave in ("FALTA", "INCAPACIDAD", "DIA_ECONOMICO", "PERMISO") else 0
            horas = 0
            minutos = 0

            # Periodo / anio / mes / quincena
            base_fecha = f_incid or f_just
            anio = mes = quincena = None
            periodo = None
            if base_fecha:
                d = datetime.strptime(base_fecha, "%Y-%m-%d")
                anio = d.year
                mes  = d.month
                quincena = 1 if d.day <= 15 else 2
                MESES_ES = ["", "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                            "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
                periodo = f"{MESES_ES[d.month]} {d.year}"

            if dry_run:
                print(f"  [DRY] fila {fila_ex:>4}  pid={pid}  tipo={tipo_clave}  "
                      f"f_incid={f_incid}  f_just={f_just}  periodo={periodo}")
                stats["dry_ok"] += 1
                continue

            try:
                with conn.cursor() as cins:
                    cins.execute(
                        """INSERT INTO incidencias
                            (personal_id, tipo_incidencia_id, estatus_id,
                             fecha_incidencia, fecha_inicio, fecha_fin, fecha_recibido_coord,
                             periodo, anio, mes, quincena,
                             dias, horas, minutos,
                             motivo, fuente)
                           VALUES (%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s,%s)
                           ON CONFLICT (personal_id, tipo_incidencia_id,
                                        COALESCE(fecha_incidencia, DATE '1900-01-01'),
                                        COALESCE(fecha_inicio,     DATE '1900-01-01'),
                                        COALESCE(fecha_fin,        DATE '1900-01-01'))
                           DO NOTHING
                           RETURNING id""",
                        (pid, tipo_id, est_registrada,
                         f_incid, f_just, f_just, f_recib,
                         periodo, anio, mes, quincena,
                         dias, horas, minutos,
                         raw_motivo, f"BASE GENERAL.xlsx!INCIDENCIAS f{fila_ex}"),
                    )
                    new = cins.fetchone()
                    if new:
                        vid = new[0]
                        stats["creados"] += 1
                        accion = "CREADO"
                    else:
                        stats["duplicados"] += 1
                        accion = "OMITIDO_DUP"
                        vid = None

                    cins.execute(
                        """UPDATE import_incidencias_raw
                              SET procesado=TRUE, incidencia_id=%s,
                                  personal_id=%s, accion=%s,
                                  fecha_incid_parsed=%s, fecha_inicio_parsed=%s,
                                  fecha_fin_parsed=%s, fecha_recib_parsed=%s,
                                  tipo_detectado=%s,
                                  estado_revision='OK', error_importacion=NULL
                            WHERE id=%s""",
                        (vid, pid, accion, f_incid, f_just, f_just, f_recib, tipo_clave, sid),
                    )
                conn.commit()

            except psycopg2.Error as pe:
                conn.rollback()
                stats["errores"] += 1
                with conn.cursor() as cerr:
                    cerr.execute(
                        """UPDATE import_incidencias_raw
                              SET estado_revision='ERROR', error_importacion=%s,
                                  personal_id=%s, tipo_detectado=%s
                            WHERE id=%s""",
                        (str(pe), pid, tipo_clave, sid),
                    )
                    conn.commit()
                print(f"  [ERR fila {fila_ex}] {pe}")

        except Exception as e:
            conn.rollback()
            stats["errores"] += 1
            with conn.cursor() as cerr:
                cerr.execute(
                    """UPDATE import_incidencias_raw
                          SET estado_revision='ERROR', error_importacion=%s
                        WHERE id=%s""",
                    (str(e), sid),
                )
                conn.commit()
            print(f"  [ERR fila {fila_ex}] {e}")

    print("\n--- Resultados incidencias ---")
    for k in ("total", "creados", "duplicados", "sin_tipo", "tipo_desconocido",
              "personal_no_encontrado", "sin_fecha", "errores", "dry_ok"):
        print(f"  {k:>25} = {stats[k]}")


def _resolver_personal(conn, num_empleado: Optional[str], nombre: Optional[str]) -> Optional[int]:
    """
    Resolucion:
      1. match exacto por numero_empleado (normalizado).
      2. match por nombre_normalizado.
      3. Si >1 match por nombre => None (ambiguedad -> revision).
      4. NO crea personal.
    """
    if not num_empleado and not nombre:
        return None

    with conn.cursor() as c:
        if num_empleado:
            c.execute(
                "SELECT id FROM personal WHERE numero_empleado = %s LIMIT 2",
                (num_empleado,),
            )
            rows = c.fetchall()
            if len(rows) == 1:
                return rows[0][0]
            # Si hay >1 por num (no deberia), caer a nombre

        if nombre:
            norm_nombre = sa(nombre)
            c.execute(
                "SELECT id FROM personal WHERE nombre_normalizado = %s LIMIT 2",
                (norm_nombre,),
            )
            rows = c.fetchall()
            if len(rows) == 1:
                return rows[0][0]
            if len(rows) > 1:
                return None  # ambigüedad

            # Fallback: nombre desordenado (apellidos<->nombres) usando tokens.
            # Comparamos set de palabras; si coincide al menos 3 tokens con uno
            # solo de personal -> match.
            tokens_excel = set(norm_nombre.split())
            if len(tokens_excel) >= 3:
                c.execute(
                    """SELECT id, nombre_normalizado FROM personal
                        WHERE nombre_normalizado ILIKE %s OR nombre_normalizado ILIKE %s""",
                    (f"%{list(tokens_excel)[0]}%", f"%{list(tokens_excel)[1]}%"),
                )
                candidatos = []
                for rid, rnom in c.fetchall():
                    tokens_db = set((rnom or "").split())
                    inter = tokens_excel & tokens_db
                    # Tolerancia ortografica ligera por 1 letra de diferencia
                    if len(inter) >= 3 or (
                        len(inter) >= 2 and
                        _match_con_tolerancia(tokens_excel - inter, tokens_db - inter)
                    ):
                        candidatos.append(rid)
                if len(candidatos) == 1:
                    return candidatos[0]

    return None


def _match_con_tolerancia(set_a: set, set_b: set) -> bool:
    """Devuelve True si al menos 1 par (a,b) difiere en maximo 1 caracter (typo)."""
    for a in set_a:
        for b in set_b:
            if abs(len(a) - len(b)) > 1:
                continue
            # distancia de Hamming aproximada cuando tienen misma longitud
            if len(a) == len(b):
                diffs = sum(1 for x, y in zip(a, b) if x != y)
                if diffs <= 1:
                    return True
            else:
                # longitud difiere en 1 -> posible insercion
                corto, largo = (a, b) if len(a) < len(b) else (b, a)
                for i in range(len(largo)):
                    if largo[:i] + largo[i+1:] == corto:
                        return True
    return False


def _marcar_revision(conn, sid: int, pid: Optional[int],
                     f_incid: Optional[str], f_ini: Optional[str],
                     f_fin: Optional[str], f_recib: Optional[str],
                     mensaje: str, tipo_clave: Optional[str]) -> None:
    with conn.cursor() as c:
        c.execute(
            """UPDATE import_incidencias_raw
                  SET estado_revision='PENDIENTE_REVISION',
                      error_importacion=%s,
                      personal_id=%s,
                      fecha_incid_parsed=%s,
                      fecha_inicio_parsed=%s,
                      fecha_fin_parsed=%s,
                      fecha_recib_parsed=%s,
                      tipo_detectado=%s
                WHERE id=%s""",
            (mensaje, pid, f_incid, f_ini, f_fin, f_recib, tipo_clave, sid),
        )
        conn.commit()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--staging",  action="store_true", help="Cargar Excel -> staging (TRUNCATE + INSERT).")
    ap.add_argument("--dry-run",  action="store_true", help="Procesar staging SIN escribir en tabla final.")
    ap.add_argument("--importar", action="store_true", help="Procesar staging -> tabla final.")
    args = ap.parse_args()

    if not (args.staging or args.dry_run or args.importar):
        ap.print_help(); sys.exit(0)

    conn = pg_connect()
    try:
        if args.staging:  cargar_staging(conn)
        if args.dry_run:  procesar(conn, dry_run=True)
        if args.importar: procesar(conn, dry_run=False)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
