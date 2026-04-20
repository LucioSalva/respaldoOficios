"""
Importador de OFICIOS DE CONOCIMIENTO.

Fuente real detectada:
  BASE GENERAL.xlsx :: hoja 'OFICIOS DE CONOCIMIENTO'
  Columnas (cabecera en fila 1):
    [0]  CON.                      (consecutivo, fila_excel)
    [1]  FECHA DE RECIBIDO         (datetime)
    [2]  NOMBRE DIRECCION          (string; match con dependencias)
    [3]  FOLIO DE QUIEN LO ENVIA   (string; folio_direccion)
    [4]  ASUNTO                    (string)
    [5]  FOLIO DE OFICIO TICS      (vacio en todo el set actual)
    [6]  REALIZO                   (vacio en todo el set actual)
    [7]  ACUSE RECIBIDO            (datetime; fecha_acuse)
    [8]  STATUS                    (string; 'ARCHIVADO' en todos)

Uso:
    python tools/importar_conocimiento.py --staging
    python tools/importar_conocimiento.py --dry-run
    python tools/importar_conocimiento.py --importar

Reglas:
  - Tipo = CONOCIMIENTO (no requiere respuesta).
  - Dependencia se matchea por nombre normalizado (sin acentos, upper,
    espacios colapsados). Si no matchea, dependencia queda NULL y la fila
    queda en estado_revision = PENDIENTE_REVISION con el error descriptivo.
  - Estado: STATUS='ARCHIVADO' -> estado 'Archivado'; cualquier otro o
    vacio -> 'De Conocimiento'.
  - Filas completamente vacias (sin fecha, sin asunto, sin direccion)
    -> estado_revision = OK con accion='OMITIDO_VACIA' (no crea oficio).
  - Duplicados: UNIQUE logico por (folio_direccion normalizado,
    fecha_recepcion, asunto normalizado). Si ya existe un oficio
    CONOCIMIENTO con esa tripleta, accion='OMITIDO_DUP'.
  - Idempotente: --staging hace TRUNCATE de import_oficios_conocimiento_raw
    ANTES de insertar. --importar solo opera filas procesado=FALSE.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import unicodedata
from datetime import datetime, date
from typing import Any, Optional

try:
    import psycopg2
    from psycopg2.extras import execute_batch, Json
except ImportError:
    print("ERROR: falta psycopg2. pip install psycopg2-binary"); sys.exit(1)
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: falta openpyxl. pip install openpyxl"); sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "BASE GENERAL.xlsx")
SHEET = "OFICIOS DE CONOCIMIENTO"
FUENTE = f"BASE GENERAL.xlsx!{SHEET}"


# ---------------- helpers -----------------

def pg_connect():
    return psycopg2.connect(
        host     = os.environ.get("PGHOST",     "localhost"),
        port     = int(os.environ.get("PGPORT", "5432")),
        dbname   = os.environ.get("PGDATABASE", "respaldooficios"),
        user     = os.environ.get("PGUSER",     "postgres"),
        password = os.environ.get("PGPASSWORD", "admin"),
    )


def sa(s: Optional[Any]) -> str:
    """Normaliza string: quita acentos, upper, colapsa espacios."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split()).strip()


def norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    if s.upper() in ("#N/A", "#N/D", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#DIV/0!"):
        return None
    return s


def parse_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def row_is_empty(row: dict) -> bool:
    """Fila basura si no tiene fecha, ni asunto, ni dependencia."""
    return (row.get("fecha_recibido") is None
            and not row.get("asunto")
            and not row.get("nombre_direccion"))


# ---------------- staging -----------------

def cargar_staging(conn, verbose: bool = True) -> int:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET}' no existe en {XLSX}", file=sys.stderr)
        return 0
    ws = wb[SHEET]

    # Cabecera en fila 1 (confirmado en inspeccion). Datos desde fila 2.
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r is None:
            continue
        # skip si todo None
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        rows.append((i, r))

    cur = conn.cursor()
    # Idempotencia dura: --staging siempre TRUNCATE su propia tabla
    cur.execute("TRUNCATE TABLE import_oficios_conocimiento_raw RESTART IDENTITY CASCADE")

    batch = []
    for fila, r in rows:
        c = lambda idx: (r[idx] if idx < len(r) else None)
        raw = {
            "fila_excel":          fila,
            "raw_cons":            norm(c(0)),
            "raw_fecha_recibido":  norm(c(1)),
            "raw_nombre_direccion": norm(c(2)),
            "raw_folio_envia":     norm(c(3)),
            "raw_asunto":          norm(c(4)),
            "raw_folio_tics":      norm(c(5)),
            "raw_realizo":         norm(c(6)),
            "raw_acuse_recibido":  norm(c(7)),
            "raw_status":          norm(c(8)),
        }
        fecha_rec_parsed = parse_date(c(1))
        fecha_acu_parsed = parse_date(c(7))

        batch.append((
            FUENTE,
            raw["fila_excel"],
            raw["raw_cons"],
            raw["raw_fecha_recibido"],
            raw["raw_nombre_direccion"],
            raw["raw_folio_envia"],
            raw["raw_asunto"],
            raw["raw_folio_tics"],
            raw["raw_realizo"],
            raw["raw_acuse_recibido"],
            raw["raw_status"],
            Json(raw),
            fecha_rec_parsed,
            fecha_acu_parsed,
        ))

    execute_batch(cur, """
        INSERT INTO import_oficios_conocimiento_raw (
            fuente, fila_excel,
            raw_cons, raw_fecha_recibido, raw_nombre_direccion,
            raw_folio_envia, raw_asunto, raw_folio_tics,
            raw_realizo, raw_acuse_recibido, raw_status,
            raw_data,
            fecha_recibido_parsed, fecha_acuse_parsed
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, batch, page_size=200)
    conn.commit()
    if verbose:
        print(f"[STAGING] Insertadas {len(batch)} filas en import_oficios_conocimiento_raw")
    return len(batch)


# ---------------- importacion real -----------------

def importar(conn, dry_run: bool = False, verbose: bool = True) -> dict:
    cur = conn.cursor()

    # Tipo CONOCIMIENTO
    cur.execute("SELECT id FROM tipos_oficio WHERE clave = 'CONOCIMIENTO'")
    row = cur.fetchone()
    if not row:
        raise RuntimeError("Tipo 'CONOCIMIENTO' no existe. Ejecuta sql/08_conocimiento.sql primero.")
    tipo_conocimiento_id = row[0]

    # Estados
    cur.execute("SELECT id, nombre FROM estados_oficio")
    estados_by_name = {sa(n): eid for (eid, n) in cur.fetchall()}
    id_archivado       = estados_by_name.get("ARCHIVADO")
    id_de_conocimiento = estados_by_name.get("DE CONOCIMIENTO")
    if id_de_conocimiento is None or id_archivado is None:
        raise RuntimeError("Estados 'Archivado' o 'De Conocimiento' no existen en estados_oficio.")

    # Dependencias: nombre normalizado -> id
    cur.execute("SELECT id, nombre FROM dependencias WHERE activo=TRUE")
    deps_map = {sa(n): did for (did, n) in cur.fetchall()}

    # Max numero_folio rango CONOCIMIENTO (20000+)
    cur.execute("SELECT COALESCE(MAX(numero_folio), 19999) FROM oficios WHERE tipo_oficio_id = %s",
                (tipo_conocimiento_id,))
    next_numero = max(int(cur.fetchone()[0]), 19999) + 1

    # Filas staging a procesar
    cur.execute("""
        SELECT id, fila_excel, raw_nombre_direccion, raw_folio_envia, raw_asunto,
               raw_status, fecha_recibido_parsed, fecha_acuse_parsed
          FROM import_oficios_conocimiento_raw
         WHERE procesado = FALSE
         ORDER BY fila_excel
    """)
    filas = cur.fetchall()

    stats = {
        "leidas":              len(filas),
        "creados":             0,
        "omitidos_vacia":      0,
        "omitidos_duplicados": 0,
        "pendientes_revision": 0,
        "errores":             0,
    }

    for (stg_id, fila, nombre_dir, folio_env, asunto,
         raw_status, fecha_rec, fecha_acu) in filas:

        # Fila vacia
        if not fecha_rec and not asunto and not nombre_dir:
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET procesado=TRUE, accion='OMITIDO_VACIA',
                       estado_revision='OK'
                 WHERE id=%s
            """, (stg_id,))
            stats["omitidos_vacia"] += 1
            continue

        # Validacion minima: fecha_recepcion y asunto
        if not fecha_rec:
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET procesado=FALSE, accion='ERROR',
                       estado_revision='PENDIENTE_REVISION',
                       error_importacion='Fecha de recibido faltante o invalida'
                 WHERE id=%s
            """, (stg_id,))
            stats["pendientes_revision"] += 1
            continue
        if not asunto:
            # Asunto es NOT NULL en la tabla. Autorellenar con placeholder
            # pero dejar PENDIENTE_REVISION para revision humana.
            asunto = f"[SIN ASUNTO - fila {fila}]"
            pendiente_por_asunto = True
        else:
            pendiente_por_asunto = False

        # Match de dependencia
        dep_id = None
        error_dep = None
        if nombre_dir:
            key = sa(nombre_dir)
            # match directo
            dep_id = deps_map.get(key)
            if dep_id is None:
                # sub-match: buscar si la dependencia esta contenida (ej "DIRECCION DE BIENESTAR"
                # matchea con "DIRECCION DE BIENESTAR, PARTICIPACION CIUDADANA Y TERRITORIAL")
                for dep_key, dep_id_candidate in deps_map.items():
                    if key and (key in dep_key or dep_key in key):
                        dep_id = dep_id_candidate
                        break
            if dep_id is None:
                error_dep = f"Dependencia no reconocida: '{nombre_dir}'"

        # Estado
        if raw_status and sa(raw_status) == "ARCHIVADO":
            estado_id = id_archivado
        else:
            estado_id = id_de_conocimiento

        # Duplicado: (folio_direccion, fecha_recepcion, asunto_normalizado, tipo)
        cur.execute("""
            SELECT id FROM oficios
             WHERE tipo_oficio_id = %s
               AND fecha_recepcion = %s
               AND COALESCE(folio_direccion,'') = COALESCE(%s,'')
               AND UPPER(asunto) = UPPER(%s)
             LIMIT 1
        """, (tipo_conocimiento_id, fecha_rec, folio_env or '', asunto))
        dup = cur.fetchone()
        if dup:
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET procesado=TRUE, accion='OMITIDO_DUP',
                       oficio_id=%s, estado_revision='OK',
                       error_importacion=NULL
                 WHERE id=%s
            """, (dup[0], stg_id))
            stats["omitidos_duplicados"] += 1
            continue

        if dry_run:
            # No hacemos INSERT en oficios, pero igual actualizamos match previews
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET dependencia_id_match=%s,
                       estado_id_match=%s,
                       estado_revision = CASE WHEN %s IS NOT NULL OR %s = TRUE
                                              THEN 'PENDIENTE_REVISION' ELSE 'OK' END,
                       error_importacion = COALESCE(%s, error_importacion)
                 WHERE id=%s
            """, (dep_id, estado_id, error_dep, pendiente_por_asunto, error_dep, stg_id))
            if error_dep or pendiente_por_asunto:
                stats["pendientes_revision"] += 1
            continue

        # INSERT oficio real
        try:
            numero_folio = next_numero
            next_numero += 1
            # fecha_rec puede venir como date (psycopg2) o str (pruebas)
            if isinstance(fecha_rec, (datetime, date)):
                anio_folio = fecha_rec.year
            else:
                anio_folio = datetime.strptime(str(fecha_rec), "%Y-%m-%d").year

            cur.execute("""
                INSERT INTO oficios
                    (numero_folio, anio_folio,
                     folio_direccion,
                     tipo_oficio_id, dependencia_id,
                     estado_id,
                     asunto, fecha_recepcion, fecha_acuse)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                numero_folio, anio_folio,
                folio_env,
                tipo_conocimiento_id, dep_id,
                estado_id,
                asunto, fecha_rec, fecha_acu,
            ))
            new_id = cur.fetchone()[0]

            # Movimiento inicial
            cur.execute("""
                INSERT INTO movimientos_oficio
                    (oficio_id, estado_anterior_id, estado_nuevo_id, observacion)
                VALUES (%s, NULL, %s, %s)
            """, (new_id, estado_id, 'Importado desde Excel OFICIOS DE CONOCIMIENTO'))

            # Actualizar staging
            revision = 'OK'
            if error_dep or pendiente_por_asunto:
                revision = 'PENDIENTE_REVISION'
                stats["pendientes_revision"] += 1
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET procesado=TRUE, accion='CREADO',
                       oficio_id=%s, dependencia_id_match=%s,
                       estado_id_match=%s,
                       estado_revision=%s,
                       error_importacion=%s
                 WHERE id=%s
            """, (new_id, dep_id, estado_id, revision, error_dep, stg_id))
            stats["creados"] += 1

        except Exception as ex:
            conn.rollback()
            cur.execute("""
                UPDATE import_oficios_conocimiento_raw
                   SET procesado=FALSE, accion='ERROR',
                       estado_revision='ERROR',
                       error_importacion=%s
                 WHERE id=%s
            """, (str(ex)[:500], stg_id))
            conn.commit()
            stats["errores"] += 1
            continue

    if not dry_run:
        conn.commit()
    else:
        conn.commit()  # solo actualiza staging con previews

    if verbose:
        print("\n=== RESUMEN IMPORTACION ===")
        for k, v in stats.items():
            print(f"  {k:22s} : {v}")
    return stats


# ---------------- CLI -----------------

def main():
    parser = argparse.ArgumentParser(description="Importador de OFICIOS DE CONOCIMIENTO")
    parser.add_argument("--staging", action="store_true", help="Cargar hoja a staging (TRUNCATE + INSERT)")
    parser.add_argument("--importar", action="store_true", help="Importar staging -> tabla oficios")
    parser.add_argument("--dry-run",  action="store_true", help="Simulacion: no crea oficios, solo reporta")
    args = parser.parse_args()

    if not (args.staging or args.importar or args.dry_run):
        parser.print_help()
        sys.exit(1)

    conn = pg_connect()
    try:
        if args.staging:
            cargar_staging(conn)
        if args.importar or args.dry_run:
            importar(conn, dry_run=args.dry_run)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
